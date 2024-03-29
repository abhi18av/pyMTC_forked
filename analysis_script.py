import xmltodict
import subprocess
from os import mkdir, chdir, remove, listdir, rename, getcwd, system
from Bio import Entrez

Entrez.email = "christophe.guyeux@univ-fcomte.fr"

REP = '/home/christophe/Documents/codes/66.MTBC_old_DNA/'
REP = '/home/christophe/Documents/codes/66.MTBC_Kinshasa/'
REP = '/run/media/christophe/76005607-4d83-44fd-9550-4027f19e822b/'
REP = '/run/media/christophe/b02a1067-f287-4665-a4c2-9129853c0b26/'
REP = '/media/christophe/MTBC/'


def entrez_to_dico(dico, loc='', Strain=''):
    dico_afr = {}
    # On ne gère (pour l'instant ?) que les ILLUMINA paired end
    if 'ILLUMINA' in dico['EXPERIMENT_PACKAGE_SET']['EXPERIMENT_PACKAGE']['EXPERIMENT']['PLATFORM'] \
            and 'PAIRED' in \
            dico['EXPERIMENT_PACKAGE_SET']['EXPERIMENT_PACKAGE']['EXPERIMENT']['DESIGN']['LIBRARY_DESCRIPTOR'][
                'LIBRARY_LAYOUT']:
        # On récupère diverses informations
        try:
            attributes = dico['EXPERIMENT_PACKAGE_SET']['EXPERIMENT_PACKAGE']['SAMPLE']['SAMPLE_ATTRIBUTES'][
                'SAMPLE_ATTRIBUTE']
        except:
            return {}
        location, date, sra, center, strain = '', '', '', '', ''
        for k in attributes:
            if k['TAG'] == 'geographic location (country and/or sea)':
                location = k['VALUE']
            elif k['TAG'] == 'collection date':
                date = k['VALUE']
            elif k['TAG'] == 'SRA accession':
                sra = k['VALUE']
            elif k['TAG'] == 'INSDC center name':
                center = k['VALUE']
            elif k['TAG'] == 'Strain':
                strain = k['VALUE']
        if loc != '':
            location = loc
        if Strain != '':
            strain = Strain
        # accession des listes d'expérimentation
        SRR = dico['EXPERIMENT_PACKAGE_SET']['EXPERIMENT_PACKAGE']['RUN_SET']['RUN']
        if not isinstance(SRR, list):
            SRR = [SRR['@accession']]
        else:
            SRR = [u['@accession'] for u in dico['EXPERIMENT_PACKAGE_SET']['EXPERIMENT_PACKAGE']['RUN_SET']['RUN']]
        for srr in SRR:
            if srr not in dico_afr:
                dico_afr[srr] = {'accession': srr,
                                 'location': location,
                                 'date': date,
                                 'SRA': sra,
                                 'center': center,
                                 'strain': strain,
                                 'SIT': '',
                                 'spoligo': '',
                                 'spoligo_new': '',
                                 'lineage_Coll': '',
                                 # 'IS_mapper': '',
                                 'IS6110': ''
                                 }
                dico_afr[srr]['taxid'] = dico['EXPERIMENT_PACKAGE_SET']['EXPERIMENT_PACKAGE']['SAMPLE']['SAMPLE_NAME'][
                    'TAXON_ID']
                dico_afr[srr]['name'] = dico['EXPERIMENT_PACKAGE_SET']['EXPERIMENT_PACKAGE']['SAMPLE']['SAMPLE_NAME'][
                    'SCIENTIFIC_NAME']

                """
                # Ajout des MIRU
                for mir in ['MIRU02', 'Mtub04', 'ETRC', 'MIRU04', 'MIRU40', 'MIRU10', 'MIRU16', 'Mtub21', 'MIRU20', 'QUB11b', 
                            'ETRA', 'Mtub29', 'Mtub30', 'ETRB', 'MIRU23', 'MIRU24', 'MIRU26', 'MIRU27', 'Mtub34', 'MIRU31', 
                            'Mtub39', 'QUB26', 'QUB4156', 'MIRU39']:
                    dico_afr[srr][mir] = ''
                """
                '''
                dico_afr[srr]['title'] = dico['EXPERIMENT_PACKAGE_SET']['EXPERIMENT_PACKAGE']['STUDY']['DESCRIPTOR']['STUDY_TITLE']
                desc = '' 
                for k in ['STUDY_ABSTRACT', 'STUDY_DESCRIPTION']:
                    if k in dico['EXPERIMENT_PACKAGE_SET']['EXPERIMENT_PACKAGE']['STUDY']['DESCRIPTOR']:
                        txt = dico['EXPERIMENT_PACKAGE_SET']['EXPERIMENT_PACKAGE']['STUDY']['DESCRIPTOR'][k]
                        if txt != desc:
                            desc = txt
                dico_afr[srr]['desc'] = desc
                '''
    return dico_afr


dico_afr = {}

from pickle import load, dump
from openpyxl import load_workbook
from xlrd import open_workbook

fichier = 'dico_africanum.pkl'
with open('data/' + fichier, 'rb') as f:
    dico_afr = load(f)

len(dico_afr)

# def entrez_to_dico(dico, loc='', Strain=''):
#     dico_afr = {}
#     # On ne gère (pour l'instant ?) que les ILLUMINA paired end
#     if 'ILLUMINA' in dico['EXPERIMENT_PACKAGE_SET']['EXPERIMENT_PACKAGE']['EXPERIMENT']['PLATFORM'] \
#             and 'PAIRED' in \
#             dico['EXPERIMENT_PACKAGE_SET']['EXPERIMENT_PACKAGE']['EXPERIMENT']['DESIGN']['LIBRARY_DESCRIPTOR'][
#                 'LIBRARY_LAYOUT']:
#         # On récupère diverses informations
#         try:
#             attributes = dico['EXPERIMENT_PACKAGE_SET']['EXPERIMENT_PACKAGE']['SAMPLE']['SAMPLE_ATTRIBUTES'][
#                 'SAMPLE_ATTRIBUTE']
#         except:
#             return {}
#         location, date, sra, center, strain = '', '', '', '', ''
#         for k in attributes:
#             if k['TAG'] == 'geographic location (country and/or sea)':
#                 location = k['VALUE']
#             elif k['TAG'] == 'collection date':
#                 date = k['VALUE']
#             elif k['TAG'] == 'SRA accession':
#                 sra = k['VALUE']
#             elif k['TAG'] == 'INSDC center name':
#                 center = k['VALUE']
#             elif k['TAG'] == 'Strain':
#                 strain = k['VALUE']
#         if loc != '':
#             location = loc
#         if Strain != '':
#             strain = Strain
#         # accession des listes d'expérimentation
#         SRR = dico['EXPERIMENT_PACKAGE_SET']['EXPERIMENT_PACKAGE']['RUN_SET']['RUN']
#         if not isinstance(SRR, list):
#             SRR = [SRR['@accession']]
#         else:
#             SRR = [u['@accession'] for u in dico['EXPERIMENT_PACKAGE_SET']['EXPERIMENT_PACKAGE']['RUN_SET']['RUN']]
#         for srr in SRR:
#             if srr not in dico_afr:
#                 dico_afr[srr] = {'accession': srr,
#                                  'location': location,
#                                  'date': date,
#                                  'SRA': sra,
#                                  'center': center,
#                                  'strain': strain,
#                                  'SIT': '',
#                                  'spoligo': '',
#                                  'spoligo_new': '',
#                                  'lineage_Coll': '',
#                                  # 'IS_mapper': '',
#                                  'IS6110': ''
#                                  }
#                 dico_afr[srr]['taxid'] = dico['EXPERIMENT_PACKAGE_SET']['EXPERIMENT_PACKAGE']['SAMPLE']['SAMPLE_NAME'][
#                     'TAXON_ID']
#                 dico_afr[srr]['name'] = dico['EXPERIMENT_PACKAGE_SET']['EXPERIMENT_PACKAGE']['SAMPLE']['SAMPLE_NAME'][
#                     'SCIENTIFIC_NAME']
#
#                 """
#                 # Ajout des MIRU
#                 for mir in ['MIRU02', 'Mtub04', 'ETRC', 'MIRU04', 'MIRU40', 'MIRU10', 'MIRU16', 'Mtub21', 'MIRU20', 'QUB11b',
#                             'ETRA', 'Mtub29', 'Mtub30', 'ETRB', 'MIRU23', 'MIRU24', 'MIRU26', 'MIRU27', 'Mtub34', 'MIRU31',
#                             'Mtub39', 'QUB26', 'QUB4156', 'MIRU39']:
#                     dico_afr[srr][mir] = ''
#                 """
#                 '''
#                 dico_afr[srr]['title'] = dico['EXPERIMENT_PACKAGE_SET']['EXPERIMENT_PACKAGE']['STUDY']['DESCRIPTOR']['STUDY_TITLE']
#                 desc = ''
#                 for k in ['STUDY_ABSTRACT', 'STUDY_DESCRIPTION']:
#                     if k in dico['EXPERIMENT_PACKAGE_SET']['EXPERIMENT_PACKAGE']['STUDY']['DESCRIPTOR']:
#                         txt = dico['EXPERIMENT_PACKAGE_SET']['EXPERIMENT_PACKAGE']['STUDY']['DESCRIPTOR'][k]
#                         if txt != desc:
#                             desc = txt
#                 dico_afr[srr]['desc'] = desc
#                 '''
#     return dico_afr


from xlrd import open_workbook

wb = open_workbook('data/1_3882_SORTED.xls')
ws = wb.sheet_by_index(0)
spol_sit = {}
for row in range(1, ws.nrows):
    spol, sit = ws.cell_value(row, 2).replace('n', '\u25A0').replace('o', '\u25A1'), ws.cell_value(row, 8)
    if spol not in spol_sit:
        spol_sit[spol] = sit

from openpyxl import load_workbook

wb = load_workbook(filename='data/Coll_62_SNPs.xlsx', read_only=True)
ws = wb['Feuil1']

h37Rv = open('data/NC_000962.3.fasta').read()
h37Rv = ''.join(h37Rv.split('\n')[1:])

longueur = 20


def change(x):
    if x == 'A':
        return 'T'
    elif x == 'T':
        return 'A'
    elif x == 'C':
        return 'G'
    elif x == 'G':
        return 'C'
    return x


def rev_comp(s):
    u = [change(x) for x in s]
    u.reverse()
    return ''.join(u)


import csv
import shutil
from Bio import pairwise2, SeqIO
from Bio.pairwise2 import format_alignment


def similaire(x, y):
    alignments = pairwise2.align.globalxx(x, y)
    return alignments[0][2] / alignments[0][4]


from pickle import dump
from datetime import datetime


def save_dico(fic='data/dico_africanum.pkl'):
    with open(fic, 'wb') as f:
        dump(dico_afr, f)
    now = datetime.now()
    with open('data/archives/dico_afr-' + now.strftime("%Y_%m_%d-%H") + '.pkl', 'wb') as f:
        dump(dico_afr, f)


# taille_amorces = 45
# dd = open('data/spoligo_old.fasta').read()
# ee = open('data/spoligo_vitro.fasta','w')
# for k in dd.split('\n'):
#     if '>' in k:
#         ee.write(k.replace('espaceur','espaceur_vitro')+'\n')
#     else:
#         txt = 'CCGAGAGGGGACGGAAAC'+k
#         ee.write(txt[:47]+'\n')
#
# for k in dd.split('\n'):
#     if '>' in k:
#         ee.write(k.replace('espaceur','espaceur_vitroB')+'\n')
#     else:
#         txt = k+'GTCGTCAGACCCAAAACC'
#         ee.write(txt[:47]+'\n')
#
# ee.close()
# dd = open('data/spoligo_new.fasta').read()
# ee = open('data/spoligo_vitro_new.fasta','w')
# for k in dd.split('\n'):
#     if '>' in k:
#         ee.write(k.replace('espaceurNew','espaceur_vitro_new')+'\n')
#     else:
#         txt = 'CCGAGAGGGGACGGAAAC'+k
#         ee.write(txt[:47]+'\n')
#
# for k in dd.split('\n'):
#     if '>' in k:
#         ee.write(k.replace('espaceurNew','espaceur_vitro_newB')+'\n')
#     else:
#         txt = k+'GTCGTCAGACCCAAAACC'
#         ee.write(txt[:47]+'\n')
#
# ee.close()

Lignee_Pali = {}
wb3 = load_workbook(filename='data/Palittapon_SNPs.xlsx', read_only=True)
fiche = wb3['Feuil1']
for row in fiche.iter_rows(min_row=2):
    if row[1].value != None:
        lignee = row[0].value.lstrip('lineage')
        pos0 = row[1].value
        source = row[3].value[0]
        cible = row[3].value[2]
        pos = pos0 - 1
        assert h37Rv[pos] == source
        seq1 = h37Rv[pos - longueur:pos + longueur + 1]
        seq2 = seq1[:20] + cible + seq1[21:]
        Lignee_Pali[pos] = (seq1, seq2, lignee)

Lignee_Shitikov = {}
wb4 = load_workbook(filename='data/Shitikov_L2_SNPs.xlsx', read_only=True)
fiche = wb4['Feuil1']
for row in fiche.iter_rows(min_row=2):
    if row[1].value != None:
        lignee = row[0].value.lstrip('lineage').replace('*', '')
        pos0 = row[1].value
        source = row[3].value[0]
        cible = row[3].value[2]
        pos = pos0 - 1
        assert h37Rv[pos] == source
        seq1 = h37Rv[pos - longueur:pos + longueur + 1]
        seq2 = seq1[:20] + cible + seq1[21:]
        Lignee_Shitikov[pos] = (seq1, seq2, lignee)

Lignee_Stucki = {}
wb5 = load_workbook(filename='data/Stucki_L4-SNPs.xlsx', read_only=True)
fiche = wb5['Feuil1']
for row in fiche.iter_rows(min_row=2):
    if row[1].value != None:
        lignee = row[0].value.lstrip('lineage').replace('*', '')
        pos0 = row[1].value
        source = row[3].value[0]
        cible = row[3].value[2]
        pos = pos0 - 1
        assert h37Rv[pos] == source
        seq1 = h37Rv[pos - longueur:pos + longueur + 1]
        seq2 = seq1[:20] + cible + seq1[21:]
        Lignee_Stucki[pos] = (seq1, seq2, lignee)

# todel = [  # 'spoligo', 'spoligo_nb',
#     # 'spoligo_new', 'spoligo_new_nb',
#     # 'SIT',
#     # 'spoligo_vitro', 'spoligo_vitro_nb',
#     # 'spoligo_vitro_new', 'spoligo_vitro_new_nb',
#     # 'SIT_silico',
#     'lineage_PGG',
#     'lineage_Coll',
#     'lineage_Pali',
#     'lineage_Shitikov',
#     'Lignee_Stucki'
# ]
#
# # todel = []
#
# for item in listdir(REP + 'sequences/'):
#     rep = REP + 'sequences/' + item + '/'
#     if 'dico.txt' in listdir(rep):
#         with open(rep + 'dico.txt') as f:
#             dicc = eval(f.read())
#         for k in todel:
#             if k in dicc:
#                 del dicc[k]
#         with open(rep + 'dico.txt', 'w') as f:
#             f.write(str(dicc))
#
# for u in todel:
#     for k in dict(dico_afr):
#         if u in dico_afr[k]:
#             del dico_afr[k][u]

Origines = [
    {'Source': "Requete SRA avec txid33894[Organism:exp] (M.tuberculosis variant africanum)",
     'Author': "NCBI",
     'study accession number': '',
     'run accessions': ['ERR2704812', 'ERR2704811', 'ERR2704810', 'ERR2704809', 'ERR2704808', 'SRR7496542',
                        'ERR2679282', 'ERR2679278', 'ERR2679258', 'ERR2679250', 'ERR2679246', 'ERR2383628',
                        'ERR2383627', 'ERR2383626', 'ERR2383625', 'ERR2383624', 'ERR2383623', 'ERR2383622',
                        'ERR2383621', 'ERR2383620', 'ERR2383619', 'ERR2383618', 'ERR2181458', 'SRR6046675',
                        'SRR6046695', 'SRR6046725', 'SRR6046334', 'SRR6046131', 'SRR6046235', 'SRR6045962',
                        'SRR6045604', 'SRR6045301', 'SRR6045425', 'SRR6045496', 'SRR6045027', 'SRR6044858',
                        'SRR6044939', 'SRR6045015', 'ERR019854', 'ERR019872', 'ERR017780', 'ERR017798', 'ERR026642',
                        'ERR1679637', 'SRR3647358', 'ERR1334053', 'ERR1334052', 'ERR1334051', 'ERR1334050',
                        'ERR1334049', 'ERR1215483', 'ERR1215482', 'ERR1215481', 'ERR1215480', 'ERR1215479',
                        'ERR1215478', 'ERR1215477', 'ERR1215476', 'ERR1215475', 'ERR1215474', 'ERR1215473',
                        'ERR1215472', 'ERR1215471', 'ERR1215470', 'ERR1215469', 'ERR1215468', 'ERR1215467',
                        'ERR1215466', 'ERR1215465', 'ERR1215464', 'ERR1215463', 'ERR1215462', 'ERR1215461',
                        'ERR1215460', 'ERR1203078', 'ERR1203077', 'ERR1203076', 'ERR1203075', 'ERR1203074',
                        'ERR1203073', 'ERR1203072', 'ERR1203071', 'ERR1203070', 'ERR1203069', 'ERR1203068',
                        'ERR1203067', 'ERR1203066', 'ERR1203065', 'ERR1203064', 'ERR1203063', 'ERR1203062',
                        'ERR1203061', 'ERR1203060', 'ERR1203059', 'ERR1203058', 'ERR1203057', 'ERR1203056',
                        'ERR1203055', 'ERR1203054', 'ERR1203053', 'ERR1203052', 'SRR3085279', 'ERR1082143',
                        'ERR1082142', 'ERR1082141', 'ERR1082140', 'ERR1082139', 'ERR1082138', 'ERR1082137',
                        'ERR1082136', 'ERR1082135', 'ERR1082134', 'ERR1082133', 'ERR1082132', 'ERR1082131',
                        'ERR1082130', 'ERR1082129', 'ERR1082128', 'ERR1082127', 'ERR1082126', 'ERR1082125',
                        'ERR1082124', 'ERR1082123', 'ERR1082122', 'ERR1082121', 'ERR1082120', 'ERR1082119',
                        'ERR1082118', 'ERR1082117', 'ERR1082116', 'ERR1082115', 'ERR1082114', 'ERR1082113',
                        'SRR2667443', 'ERR909754', 'ERR909753', 'ERR845916', 'ERR751349', 'ERR751348', 'ERR751347',
                        'ERR751346', 'ERR751345', 'ERR751344', 'ERR751343', 'ERR751342', 'ERR751341', 'ERR751340',
                        'ERR751339', 'ERR751338', 'ERR751337', 'ERR751336', 'ERR751335', 'ERR751334', 'ERR751333',
                        'ERR751332', 'ERR751331', 'ERR751330', 'ERR751329', 'ERR751328', 'ERR751327', 'ERR751326',
                        'ERR751325', 'ERR751324', 'ERR751323', 'ERR751322', 'ERR751321', 'ERR751320', 'ERR751319',
                        'ERR751318', 'ERR751317', 'ERR751316', 'ERR751315', 'ERR751314', 'ERR751313', 'ERR751312',
                        'ERR751311', 'ERR751310', 'ERR751309', 'ERR751308', 'ERR751307', 'ERR751306', 'ERR751305',
                        'ERR751304', 'ERR751303', 'ERR751302', 'ERR751301', 'ERR751300', 'ERR751299', 'ERR751298',
                        'ERR751297', 'ERR751296', 'ERR751295', 'ERR751294', 'ERR751293', 'ERR751292', 'ERR751291',
                        'ERR751290', 'ERR702414', 'ERR702413', 'ERR702412', 'ERR702411', 'ERR702410', 'ERR702409',
                        'ERR702408', 'ERR702407', 'ERR702406', 'ERR702405', 'ERR702404', 'ERR702403', 'ERR702402',
                        'ERR702401', 'ERR702400', 'ERR702399', 'ERR702437', 'ERR702436', 'ERR702435', 'ERR702434',
                        'ERR702433', 'ERR702432', 'ERR702431', 'ERR702430', 'ERR702429', 'ERR702428', 'ERR702427',
                        'ERR702426', 'ERR702425', 'ERR702424', 'ERR702423', 'ERR702422', 'ERR702421', 'ERR702420',
                        'ERR702419', 'ERR702418', 'ERR702417', 'ERR702416', 'ERR702415', 'ERR517472', 'ERR517471',
                        'ERR502538', 'ERR502537', 'ERR502536', 'ERR502535', 'ERR502534', 'ERR502533', 'ERR502532',
                        'ERR502531', 'ERR502530', 'ERR502529', 'ERR502528', 'ERR502527', 'ERR502526', 'ERR502525',
                        'ERR502524', 'ERR502523', 'ERR502522', 'ERR502521', 'ERR502520', 'ERR502519', 'ERR502518',
                        'ERR502517', 'ERR502516', 'ERR502515', 'ERR502514', 'ERR502513', 'ERR502512', 'ERR502511',
                        'ERR502510', 'ERR502509', 'ERR502508', 'ERR502507', 'ERR502506', 'ERR502505', 'ERR502504',
                        'ERR502503', 'ERR502502', 'ERR502501', 'ERR502500', 'ERR502499', 'ERR502498', 'ERR502497',
                        'ERR502496', 'ERR502495', 'ERR502494', 'ERR502493', 'ERR502492', 'ERR502491', 'ERR502490',
                        'ERR502489', 'ERR502488', 'ERR502487', 'ERR502486', 'ERR502485', 'ERR502484', 'ERR502483',
                        'ERR502482', 'ERR502481', 'ERR502480', 'ERR502479', 'ERR502478', 'ERR502477', 'ERR502476',
                        'ERR502475', 'ERR502474', 'ERR502473', 'ERR502472', 'ERR502471', 'ERR502470', 'SRR1735564',
                        'SRR1735577', 'SRR1735574', 'SRR1735569', 'SRR1577834', 'SRR1577833', 'SRR1577831',
                        'SRR1577828', 'SRR1577820', 'SRR1577819', 'SRR1162789', 'SRR1162788', 'SRR1162738',
                        'SRR1162716', 'SRR1162479', 'SRR1162477', 'SRR1162470', 'SRR1162469', 'SRR1103551',
                        'SRR1103499', 'SRR1103472', 'SRR1103387', 'SRR1057937', 'SRR1049960', 'SRR1049959',
                        'SRR1049958', 'SRR998743', 'SRR998740', 'SRR998741', 'SRR998742', 'SRR998655', 'SRR998652',
                        'SRR998654', 'SRR998653', 'SRR998648', 'SRR998650', 'SRR998651', 'SRR998649', 'SRR998646',
                        'SRR998647', 'SRR998645', 'SRR998644', 'SRR998640', 'SRR998641', 'SRR998642', 'SRR998643',
                        'SRR998637', 'SRR998638', 'SRR998636', 'SRR998639', 'SRR998631', 'SRR998628', 'SRR998629',
                        'SRR998630', 'SRR998626', 'SRR998624', 'SRR998625', 'SRR998627', 'SRR998621', 'SRR998620',
                        'SRR998622', 'SRR998623', 'SRR998619', 'SRR998616', 'SRR998617', 'SRR998618', 'SRR998612',
                        'SRR998614', 'SRR998613', 'SRR998615', 'SRR998611', 'SRR998608', 'SRR998610', 'SRR998609',
                        'SRR998604', 'SRR998605', 'SRR998606', 'SRR998607', 'SRR998601', 'SRR998603', 'SRR998600',
                        'SRR998602', 'SRR998596', 'SRR998597', 'SRR998598', 'SRR998599', 'SRR998594', 'SRR998592',
                        'SRR998593', 'SRR998595', 'SRR998584', 'SRR998587', 'SRR998586', 'SRR998585', 'SRR998580',
                        'SRR998583', 'SRR998582', 'SRR998581', 'SRR998578', 'SRR998576', 'SRR998579', 'SRR998577']},
    #
    {'Source': "Requete SRA avec txid78331[Organism:exp] (M.canettii)",
     'Author': "NCBI",
     'study accession number': '',
     'run accessions': ['SRR6650709', 'ERR015598', 'ERR1336826', 'ERR1336825', 'ERR1336824', 'ERR1336823', 'ERR1336822',
                        'ERR1336821', 'ERR1336820', 'ERR1043164', 'ERR1043163', 'ERR313117', 'ERR313116', 'ERR313115',
                        'ERR313114', 'ERR313113', 'ERR266126', 'ERR266125', 'ERR266124', 'ERR266123', 'ERR266122',
                        'ERR266121', 'ERR266120', 'ERR266110', 'ERR266109', 'ERR266108', 'ERR266107', 'ERR266106']},
    #
    {
        'Source': "Unexpected Genomic and Phenotypic Diversity of Mycobacterium africanum Lineage 5 Affects Drug Resistance, Protein Secretion, and Immunogenicity",
        'Author': "Ates et al. 2018",
        'study accession number': 'PRJEB25506',
        'run accessions': ['ERR2383618', 'ERR2383619', 'ERR2383620', 'ERR2383621', 'ERR2383622', 'ERR2383623',
                           'ERR2383624', 'ERR2383625', 'ERR2383626', 'ERR2383627', 'ERR2383628', 'ERR2704808',
                           'ERR2704809', 'ERR2704810', 'ERR2704812']},
    #
    {'Source': "Genomic characterization of Mycobacterium tuberculosis lineage 7 and a proposed name: 'Aethiops vetus'",
     'Author': "Nebenzahl-Guimaraes H, Yimer S, Holm-Hansen C, de Beer J, Brosch R, van Soolingen D",
     'study accession number': 'PRJEB8432',
     'run accessions': ['ERR756344', 'ERR756345', 'ERR756346', 'ERR756347', 'ERR756348']},
    #
    {
        'Source': "Analysis of IS6110 insertion sites provide a glimpse into genome evolution of Mycobacterium tuberculosis",
        'Author': "Roychowdhury T, Mandal S, Bhattacharya A",
        'study accession number': '',
        'run accessions': ['ERR015609', 'ERR015610', 'ERR015611', 'ERR015612', 'ERR015613', 'ERR015614', 'ERR015615',
                           'ERR015616', 'ERR017797', 'ERR019565', 'ERR019569', 'ERR019570', 'ERR019571', 'ERR019572',
                           'ERR019573', 'ERR019574', 'ERR019575', 'ERR023731', 'ERR023732', 'ERR023739', 'ERR023740',
                           'ERR023743', 'ERR023746', 'ERR023747', 'ERR023757', 'ERR023758', 'ERR023759', 'ERR023760',
                           'ERR023764', 'ERR024340', 'ERR024343', 'ERR024350', 'ERR024359', 'ERR025414', 'ERR025415',
                           'ERR025416', 'ERR025417', 'ERR025418', 'ERR025420', 'ERR025421', 'ERR025422', 'ERR025423',
                           'ERR025424', 'ERR025426', 'ERR025427', 'ERR025428', 'ERR025429', 'ERR025430', 'ERR025431',
                           'ERR025432', 'ERR025433', 'ERR025434', 'ERR025435', 'ERR025436', 'ERR025438', 'ERR025440',
                           'ERR025441', 'ERR025442', 'ERR025443', 'ERR025444', 'ERR025445', 'ERR025447', 'ERR025448',
                           'ERR025457', 'ERR025833', 'ERR025834', 'ERR025837', 'ERR025838', 'ERR025839', 'ERR025840',
                           'ERR025842', 'ERR025843', 'ERR025844', 'ERR025846', 'ERR025847', 'ERR025848', 'ERR026472',
                           'ERR026473', 'ERR026474', 'ERR026475', 'ERR026476', 'ERR026477', 'ERR026478', 'ERR026479',
                           'ERR026481', 'ERR026482', 'ERR027458', 'ERR027459', 'ERR027463', 'ERR027468', 'ERR028608',
                           'ERR028609', 'ERR028610', 'ERR028611', 'ERR028612', 'ERR028613', 'ERR028615', 'ERR028616',
                           'ERR028617', 'ERR028619', 'ERR028620', 'ERR028621', 'ERR028623', 'ERR028624', 'ERR028625',
                           'ERR028626', 'ERR028627', 'ERR028628', 'ERR029201', 'ERR029202', 'ERR029203', 'ERR029204',
                           'ERR029205', 'ERR029207', 'ERR029209', 'ERR029210', 'ERR029211', 'ERR036186', 'ERR036187',
                           'ERR036188', 'ERR036189', 'ERR036190', 'ERR036192', 'ERR036193', 'ERR036195', 'ERR036196',
                           'ERR036197', 'ERR036200', 'ERR036201', 'ERR036202', 'ERR036203', 'ERR036204', 'ERR036206',
                           'ERR036208', 'ERR036209', 'ERR036210', 'ERR036212', 'ERR036213', 'ERR036214', 'ERR036216',
                           'ERR036217', 'ERR036218', 'ERR036219', 'ERR036221', 'ERR036222', 'ERR036223', 'ERR036226',
                           'ERR036227', 'ERR036228', 'ERR036230', 'ERR036231', 'ERR036232', 'ERR036234', 'ERR036236',
                           'ERR036238', 'ERR036239', 'ERR036240', 'ERR036241', 'ERR036242', 'ERR036246', 'ERR036247',
                           'ERR036248', 'ERR036249', 'ERR037467', 'ERR037468', 'ERR037469', 'ERR037470', 'ERR037471',
                           'ERR037472', 'ERR037473', 'ERR037474', 'ERR037475', 'ERR037476', 'ERR037477', 'ERR037478',
                           'ERR037479', 'ERR037480', 'ERR037481', 'ERR037482', 'ERR037483', 'ERR037484', 'ERR037485',
                           'ERR037486', 'ERR037487', 'ERR037488', 'ERR037489', 'ERR037490', 'ERR037491', 'ERR037492',
                           'ERR037493', 'ERR037494', 'ERR037495', 'ERR037496', 'ERR037497', 'ERR037498', 'ERR037499',
                           'ERR037500', 'ERR037501', 'ERR037502', 'ERR037503', 'ERR037504', 'ERR037505', 'ERR037506',
                           'ERR037507', 'ERR037508', 'ERR037509', 'ERR037510', 'ERR037511', 'ERR037512', 'ERR037513',
                           'ERR037514', 'ERR037515', 'ERR037516', 'ERR037517', 'ERR037518', 'ERR037519', 'ERR037520',
                           'ERR037521', 'ERR037522', 'ERR037523', 'ERR037524', 'ERR037526', 'ERR037527', 'ERR037528',
                           'ERR037529', 'ERR037530', 'ERR037531', 'ERR037532', 'ERR037533', 'ERR037535', 'ERR037536',
                           'ERR037537', 'ERR037538', 'ERR037539', 'ERR037540', 'ERR037542', 'ERR037544', 'ERR037545',
                           'ERR037546', 'ERR037548', 'ERR037549', 'ERR037550', 'ERR037551', 'ERR037552', 'ERR037553',
                           'ERR037554', 'ERR037555', 'ERR038253', 'ERR038254', 'ERR038255', 'ERR038256', 'ERR038257',
                           'ERR038260', 'ERR038261', 'ERR038264', 'ERR038265', 'ERR038266', 'ERR038269', 'ERR038270',
                           'ERR038272', 'ERR038273', 'ERR038274', 'ERR038275', 'ERR038276', 'ERR038277', 'ERR038278',
                           'ERR038279', 'ERR038281', 'ERR038283', 'ERR038284', 'ERR038285', 'ERR038286', 'ERR038287',
                           'ERR038288', 'ERR038290', 'ERR038291', 'ERR038292', 'ERR038293', 'ERR038298', 'ERR038299',
                           'ERR038300', 'ERR038736', 'ERR038737', 'ERR038738', 'ERR038739', 'ERR038740', 'ERR038743',
                           'ERR038744', 'ERR038745', 'ERR038746', 'ERR038747', 'ERR038748', 'ERR038749', 'ERR038751',
                           'ERR038752', 'ERR038753', 'ERR038754', 'ERR038755', 'ERR039324', 'ERR039325', 'ERR039326',
                           'ERR039327', 'ERR039328', 'ERR039329', 'ERR039330', 'ERR039331', 'ERR039332', 'ERR039333',
                           'ERR039334', 'ERR039335', 'ERR039336', 'ERR039337', 'ERR039338', 'ERR039339', 'ERR039340',
                           'ERR039341', 'ERR039342', 'ERR039344', 'ERR039345', 'ERR039346', 'ERR040086', 'ERR040087',
                           'ERR040088', 'ERR040089', 'ERR040090', 'ERR040091', 'ERR040093', 'ERR040094', 'ERR040095',
                           'ERR040096', 'ERR040097', 'ERR040098', 'ERR040099', 'ERR040100', 'ERR040101', 'ERR040102',
                           'ERR040103', 'ERR040104', 'ERR040105', 'ERR040106', 'ERR040107', 'ERR040108', 'ERR040109',
                           'ERR040114', 'ERR040115', 'ERR040118', 'ERR040119', 'ERR040120', 'ERR040121', 'ERR040122',
                           'ERR040123', 'ERR040125', 'ERR040126', 'ERR040127', 'ERR040129', 'ERR040131', 'ERR040132',
                           'ERR040133', 'ERR040134', 'ERR040136', 'ERR040137', 'ERR040138', 'ERR040139', 'ERR040140',
                           'ERR040141', 'ERR040142', 'ERR046729', 'ERR046730', 'ERR046732', 'ERR046733', 'ERR046734',
                           'ERR046735', 'ERR046736', 'ERR046737', 'ERR046738', 'ERR046739', 'ERR046741', 'ERR046743',
                           'ERR046744', 'ERR046745', 'ERR046746', 'ERR046751', 'ERR046752', 'ERR046753', 'ERR046754',
                           'ERR046755', 'ERR046756', 'ERR046758', 'ERR046759', 'ERR046760', 'ERR046761', 'ERR046762',
                           'ERR046763', 'ERR046764', 'ERR046765', 'ERR046766', 'ERR046767', 'ERR046768', 'ERR046769',
                           'ERR046770', 'ERR046771', 'ERR046772', 'ERR046773', 'ERR046775', 'ERR046776', 'ERR046777',
                           'ERR046778', 'ERR046779', 'ERR046780', 'ERR046782', 'ERR046783', 'ERR046791', 'ERR046796',
                           'ERR046797', 'ERR046798', 'ERR046799', 'ERR046819', 'ERR046820', 'ERR046821', 'ERR046822',
                           'ERR046825', 'ERR046831', 'ERR046832', 'ERR046833', 'ERR046834', 'ERR046836', 'ERR046837',
                           'ERR046838', 'ERR046839', 'ERR046840', 'ERR046841', 'ERR046842', 'ERR046843', 'ERR046847',
                           'ERR046851', 'ERR046852', 'ERR046854', 'ERR046857', 'ERR046858', 'ERR046860', 'ERR046861',
                           'ERR046862', 'ERR046863', 'ERR046866', 'ERR046867', 'ERR046868', 'ERR046869', 'ERR046870',
                           'ERR046871', 'ERR046872', 'ERR046873', 'ERR046874', 'ERR046875', 'ERR046876', 'ERR046877',
                           'ERR046878', 'ERR046879', 'ERR046880', 'ERR046881', 'ERR046882', 'ERR046883', 'ERR046884',
                           'ERR046885', 'ERR046887', 'ERR046888', 'ERR046889', 'ERR046890', 'ERR046892', 'ERR046894',
                           'ERR046897', 'ERR046898', 'ERR046900', 'ERR046901', 'ERR046903', 'ERR046904', 'ERR046905',
                           'ERR046906', 'ERR046907', 'ERR046908', 'ERR046910', 'ERR046912', 'ERR046914', 'ERR046915',
                           'ERR046916', 'ERR046917', 'ERR046918', 'ERR046919', 'ERR046920', 'ERR046921', 'ERR046922',
                           'ERR046923', 'ERR046924', 'ERR046925', 'ERR046926', 'ERR046927', 'ERR046928', 'ERR046929',
                           'ERR046930', 'ERR046932', 'ERR046933', 'ERR046934', 'ERR046935', 'ERR046936', 'ERR046937',
                           'ERR046938', 'ERR046939', 'ERR046940', 'ERR046941', 'ERR046942', 'ERR046943', 'ERR046946',
                           'ERR046947', 'ERR046948', 'ERR046949', 'ERR046950', 'ERR046951', 'ERR046952', 'ERR046953',
                           'ERR046957', 'ERR046958', 'ERR046959', 'ERR046960', 'ERR046962', 'ERR046963', 'ERR046964',
                           'ERR046965', 'ERR046966', 'ERR046967', 'ERR046968', 'ERR046969', 'ERR046970', 'ERR046971',
                           'ERR046972', 'ERR046974', 'ERR046975', 'ERR046980', 'ERR046981', 'ERR046982', 'ERR046983',
                           'ERR046986', 'ERR046988', 'ERR046990', 'ERR046991', 'ERR046992', 'ERR046993', 'ERR046994',
                           'ERR046995', 'ERR046996', 'ERR046997', 'ERR046998', 'ERR046999', 'ERR047000', 'ERR047001',
                           'ERR047002', 'ERR047003', 'ERR047004', 'ERR047005', 'ERR047006', 'ERR047008', 'ERR047009',
                           'ERR047010', 'ERR047011', 'ERR047012', 'ERR047013', 'ERR047016', 'ERR047881', 'ERR047882',
                           'ERR047883', 'ERR047885', 'ERR047886', 'ERR047887', 'ERR047888', 'ERR047889', 'ERR047890',
                           'ERR047891', 'ERR067576', 'ERR067577', 'ERR067578', 'ERR067579', 'ERR067580', 'ERR067581',
                           'ERR067582', 'ERR067583', 'ERR067584', 'ERR067586', 'ERR067587', 'ERR067588', 'ERR067589',
                           'ERR067590', 'ERR067591', 'ERR067592', 'ERR067593', 'ERR067594', 'ERR067595', 'ERR067596',
                           'ERR067597', 'ERR067598', 'ERR067599', 'ERR067600', 'ERR067603', 'ERR067604', 'ERR067605',
                           'ERR067606', 'ERR067607', 'ERR067608', 'ERR067609', 'ERR067610', 'ERR067611', 'ERR067612',
                           'ERR067613', 'ERR067614', 'ERR067616', 'ERR067617', 'ERR067618', 'ERR067619', 'ERR067621',
                           'ERR067622', 'ERR067623', 'ERR067624', 'ERR067625', 'ERR067626', 'ERR067627', 'ERR067628',
                           'ERR067629', 'ERR067630', 'ERR067631', 'ERR067632', 'ERR067633', 'ERR067634', 'ERR067635',
                           'ERR067637', 'ERR067638', 'ERR067639', 'ERR067640', 'ERR067641', 'ERR067643', 'ERR067644',
                           'ERR067645', 'ERR067646', 'ERR067647', 'ERR067648', 'ERR067649', 'ERR067650', 'ERR067651',
                           'ERR067652', 'ERR067653', 'ERR067654', 'ERR067655', 'ERR067656', 'ERR067657', 'ERR067658',
                           'ERR067659', 'ERR067660', 'ERR067662', 'ERR067663', 'ERR067665', 'ERR067666', 'ERR067667',
                           'ERR067668', 'ERR067669', 'ERR067670', 'ERR067671', 'ERR067672', 'ERR067673', 'ERR067674',
                           'ERR067675', 'ERR067676', 'ERR067677', 'ERR067680', 'ERR067683', 'ERR067684', 'ERR067686',
                           'ERR067687', 'ERR067688', 'ERR067689', 'ERR067690', 'ERR067692', 'ERR067693', 'ERR067694',
                           'ERR067695', 'ERR067696', 'ERR067697', 'ERR067698', 'ERR067702', 'ERR067703', 'ERR067704',
                           'ERR067705', 'ERR067706', 'ERR067707', 'ERR067708', 'ERR067709', 'ERR067710', 'ERR067711',
                           'ERR067713', 'ERR067714', 'ERR067715', 'ERR067717', 'ERR067719', 'ERR067720', 'ERR067721',
                           'ERR067722', 'ERR067723', 'ERR067724', 'ERR067725', 'ERR067726', 'ERR067727', 'ERR067731',
                           'ERR067736', 'ERR067737', 'ERR067739', 'ERR067740', 'ERR067741', 'ERR067742', 'ERR067743',
                           'ERR067744', 'ERR067745', 'ERR067746', 'ERR067747', 'ERR067749', 'ERR067750', 'ERR067752',
                           'ERR067754', 'ERR067755', 'ERR067758', 'ERR067759', 'ERR067760', 'ERR067761', 'ERR067762',
                           'ERR067763', 'ERR067764', 'ERR067765', 'ERR067766', 'ERR072019', 'ERR072020', 'ERR072021',
                           'ERR072022', 'ERR072023', 'ERR072024', 'ERR072025', 'ERR072026', 'ERR072027', 'ERR072028',
                           'ERR072029', 'ERR072031', 'ERR072032', 'ERR072034', 'ERR072035', 'ERR072036', 'ERR072037',
                           'ERR072038', 'ERR072039', 'ERR072040', 'ERR072041', 'ERR072042', 'ERR072044', 'ERR072045',
                           'ERR072046', 'ERR072047', 'ERR072048', 'ERR072050', 'ERR072051', 'ERR072065', 'ERR072072',
                           'ERR072087', 'ERR072088', 'ERR072089', 'ERR072090', 'ERR072094', 'ERR072095', 'ERR072096',
                           'ERR192249', 'ERR192250', 'ERR233346', 'ERR233347', 'ERR233348', 'ERR233349', 'ERR233350',
                           'ERR233351', 'ERR233352', 'ERR233353', 'ERR233354', 'ERR233355', 'ERR233356', 'ERR233357',
                           'ERR233358', 'ERR233359', 'ERR233360', 'ERR233361', 'ERR233362', 'ERR233363', 'ERR233364',
                           'ERR233365', 'ERR233366', 'ERR233367', 'ERR233368', 'ERR233369', 'ERR233370', 'ERR233371',
                           'ERR233372', 'ERR233373', 'ERR233374', 'ERR233375', 'ERR233376', 'ERR233377', 'ERR233378',
                           'ERR233379', 'ERR233380', 'ERR233381', 'ERR233382', 'ERR233383', 'ERR233384', 'ERR233385',
                           'ERR233386', 'ERR233387', 'ERR233388', 'ERR233389', 'ERR233390', 'ERR233391', 'ERR234097',
                           'ERR234098', 'ERR234099', 'ERR234100', 'ERR234101', 'ERR234102', 'ERR234103', 'ERR234104',
                           'ERR234105', 'ERR234106', 'ERR234107', 'ERR234108', 'ERR234109', 'ERR234110', 'ERR234111',
                           'ERR234112', 'ERR234113', 'ERR234114', 'ERR234115', 'ERR234116', 'ERR234117', 'ERR234118',
                           'ERR234119', 'ERR234120', 'ERR234121', 'ERR234122', 'ERR234123', 'ERR234124', 'ERR234125',
                           'ERR234126', 'ERR234127', 'ERR234128', 'ERR234129', 'ERR234130', 'ERR234131', 'ERR234132',
                           'ERR234133', 'ERR234134', 'ERR234135', 'ERR234136', 'ERR234137', 'ERR234138', 'ERR234139',
                           'ERR234140', 'ERR234141', 'ERR234142', 'ERR234143', 'ERR234144', 'ERR234145', 'ERR234146',
                           'ERR234147', 'ERR234148', 'ERR234149', 'ERR234150', 'ERR234151', 'ERR234152', 'ERR234153',
                           'ERR234154', 'ERR234155', 'ERR234156', 'ERR234157', 'ERR234158', 'ERR234159', 'ERR234160',
                           'ERR234161', 'ERR234162', 'ERR234163', 'ERR234164', 'ERR234165', 'ERR234166', 'ERR234167',
                           'ERR234168', 'ERR234169', 'ERR234170', 'ERR234171', 'ERR234172', 'ERR234173', 'ERR234174',
                           'ERR234175', 'ERR234176', 'ERR234177', 'ERR234178', 'ERR234179', 'ERR234180', 'ERR234181',
                           'ERR234182', 'ERR234183', 'ERR234184', 'ERR234185', 'ERR234186', 'ERR234187', 'ERR234188',
                           'ERR234189', 'ERR234190', 'ERR234191', 'ERR234192', 'ERR234193', 'ERR234194', 'ERR234195',
                           'ERR234196', 'ERR234197', 'ERR234198', 'ERR234199', 'ERR234200', 'ERR234201', 'ERR234202',
                           'ERR234203', 'ERR234204', 'ERR234205', 'ERR234206', 'ERR234207', 'ERR234208', 'ERR234209',
                           'ERR234210', 'ERR234211', 'ERR234212', 'ERR234213', 'ERR234215', 'ERR234216', 'ERR234217',
                           'ERR234224', 'ERR234226', 'ERR234232', 'ERR234233', 'ERR234234', 'ERR234235', 'ERR234236',
                           'ERR234237', 'ERR234238', 'ERR234239', 'ERR234240', 'ERR234241', 'ERR234242', 'ERR234243',
                           'ERR234244', 'ERR234245', 'ERR234246', 'ERR234247', 'ERR234248', 'ERR234249', 'ERR234250',
                           'ERR234251', 'ERR234252', 'ERR234253', 'ERR234254', 'ERR234255', 'ERR234256', 'ERR234257',
                           'ERR234258', 'ERR234259', 'ERR234260', 'ERR234261', 'ERR234262', 'ERR234263', 'ERR234264',
                           'ERR234265', 'ERR234266', 'ERR234267', 'ERR234268', 'ERR234269', 'ERR234270', 'ERR234271',
                           'ERR234272', 'ERR234273', 'ERR238745', 'ERR238746', 'ERR245646', 'ERR245647', 'ERR245648',
                           'ERR245650', 'ERR245651', 'ERR245652', 'ERR245653', 'ERR245654', 'ERR245655', 'ERR245656',
                           'ERR245657', 'ERR245658', 'ERR245659', 'ERR245660', 'ERR245661', 'ERR245662', 'ERR245663',
                           'ERR245664', 'ERR245666', 'ERR245667', 'ERR245668', 'ERR245669', 'ERR245670', 'ERR245671',
                           'ERR245672', 'ERR245673', 'ERR245674', 'ERR245676', 'ERR245677', 'ERR245678', 'ERR245679',
                           'ERR245680', 'ERR245682', 'ERR245683', 'ERR245684', 'ERR245685', 'ERR245686', 'ERR245687',
                           'ERR245688', 'ERR245689', 'ERR245690', 'ERR245692', 'ERR245693', 'ERR245694', 'ERR245695',
                           'ERR245696', 'ERR245697', 'ERR245700', 'ERR245701', 'ERR245702', 'ERR245703', 'ERR245704',
                           'ERR245705', 'ERR245706', 'ERR245707', 'ERR245708', 'ERR245709', 'ERR245710', 'ERR245711',
                           'ERR245712', 'ERR245713', 'ERR245715', 'ERR245716', 'ERR245717', 'ERR245718', 'ERR245719',
                           'ERR245720', 'ERR245722', 'ERR245723', 'ERR245724', 'ERR245725', 'ERR245726', 'ERR245727',
                           'ERR245728', 'ERR245729', 'ERR245730', 'ERR245731', 'ERR245732', 'ERR245733', 'ERR245734',
                           'ERR245736', 'ERR245737', 'ERR245738', 'ERR245739', 'ERR245740', 'ERR245741', 'ERR245742',
                           'ERR245743', 'ERR245744', 'ERR245745', 'ERR245746', 'ERR245747', 'ERR245748', 'ERR245749',
                           'ERR245750', 'ERR245751', 'ERR245752', 'ERR245753', 'ERR245754', 'ERR245755', 'ERR245756',
                           'ERR245757', 'ERR245758', 'ERR245759', 'ERR245760', 'ERR245761', 'ERR245762', 'ERR245763',
                           'ERR245764', 'ERR245765', 'ERR245767', 'ERR245768', 'ERR245769', 'ERR245770', 'ERR245771',
                           'ERR245772', 'ERR245773', 'ERR245774', 'ERR245775', 'ERR245776', 'ERR245777', 'ERR245778',
                           'ERR245779', 'ERR245780', 'ERR245781', 'ERR245783', 'ERR245784', 'ERR245785', 'ERR245786',
                           'ERR245787', 'ERR245788', 'ERR245789', 'ERR245792', 'ERR245793', 'ERR245794', 'ERR245796',
                           'ERR245797', 'ERR245798', 'ERR245799', 'ERR245800', 'ERR245801', 'ERR245802', 'ERR245803',
                           'ERR245804', 'ERR245805', 'ERR245806', 'ERR245807', 'ERR245808', 'ERR245809', 'ERR245810',
                           'ERR245812', 'ERR245813', 'ERR245815', 'ERR245818', 'ERR245821', 'ERR245822', 'ERR245826',
                           'ERR245827', 'ERR245828', 'ERR245829', 'ERR245830', 'ERR245831', 'ERR245835', 'ERR245836',
                           'ERR245837', 'ERR245838', 'ERR245839', 'ERR245840', 'ERR245841', 'ERR245842', 'ERR245843',
                           'ERR245844', 'ERR245845', 'ERR245846', 'ERR245847', 'ERR245848', 'ERR245849', 'ERR257891',
                           'ERR257892', 'ERR257893', 'ERR257894', 'ERR257895', 'ERR257896', 'ERR257897', 'ERR257898',
                           'ERR257899', 'ERR257900', 'ERR257901', 'ERR257902', 'ERR257903', 'ERR257904', 'ERR257905',
                           'ERR257906', 'ERR257907', 'ERR257908', 'ERR257909', 'ERR257910', 'ERR257911', 'ERR257912',
                           'ERR257915', 'ERR257917', 'ERR257923', 'ERR257925', 'ERR257927', 'ERR257928', 'ERR257929',
                           'ERR257930', 'ERR257931', 'ERR257932', 'ERR257933', 'ERR257934', 'ERR257935', 'SRR006916',
                           'SRR006917', 'SRR006918', 'SRR006919', 'SRR006920', 'SRR014813', 'SRR014814', 'SRR017356',
                           'SRR022869', 'SRR022870', 'SRR022872', 'SRR022873', 'SRR022874', 'SRR022875', 'SRR022876',
                           'SRR022877', 'SRR022878', 'SRR022879', 'SRR022880', 'SRR029156', 'SRR671719', 'SRR671720',
                           'SRR671721', 'SRR671722', 'SRR671723', 'SRR671724', 'SRR671725', 'SRR671726', 'SRR671727',
                           'SRR671728', 'SRR671729', 'SRR671730', 'SRR671731', 'SRR671732', 'SRR671733', 'SRR671734',
                           'SRR671735', 'SRR671736', 'SRR671737', 'SRR671738', 'SRR671739', 'SRR671740', 'SRR671741',
                           'SRR671742', 'SRR671743', 'SRR671744', 'SRR671745', 'SRR671746', 'SRR671747', 'SRR671748',
                           'SRR671749', 'SRR671750', 'SRR671751', 'SRR671752', 'SRR671753', 'SRR671754', 'SRR671755',
                           'SRR671756', 'SRR671757', 'SRR671758', 'SRR671759', 'SRR671760', 'SRR671761', 'SRR671762',
                           'SRR671763', 'SRR671764', 'SRR671765', 'SRR671766', 'SRR671767', 'SRR671768', 'SRR671769',
                           'SRR671770', 'SRR671771', 'SRR671772', 'SRR671773', 'SRR671774', 'SRR671775', 'SRR671776',
                           'SRR671777', 'SRR671778', 'SRR671779', 'SRR671780', 'SRR671781', 'SRR671782', 'SRR671783',
                           'SRR671784', 'SRR671785', 'SRR671786', 'SRR671787', 'SRR671788', 'SRR671789', 'SRR671790',
                           'SRR671791', 'SRR671792', 'SRR671793', 'SRR671794', 'SRR671795', 'SRR671796', 'SRR671797',
                           'SRR671798', 'SRR671799', 'SRR671800', 'SRR671801', 'SRR671802', 'SRR671803', 'SRR671804',
                           'SRR671805', 'SRR671806', 'SRR671807', 'SRR671808', 'SRR671809', 'SRR671810', 'SRR671811',
                           'SRR671812', 'SRR671813', 'SRR671814', 'SRR671815', 'SRR671816', 'SRR671817', 'SRR671818',
                           'SRR671819', 'SRR671820', 'SRR671821', 'SRR671822', 'SRR671823', 'SRR671824', 'SRR671825',
                           'SRR671826', 'SRR671827', 'SRR671828', 'SRR671829', 'SRR671830', 'SRR671831', 'SRR671832',
                           'SRR671833', 'SRR671834', 'SRR671835', 'SRR671836', 'SRR671837', 'SRR671838', 'SRR671839',
                           'SRR671840', 'SRR671841', 'SRR671842', 'SRR671843', 'SRR671844', 'SRR671845', 'SRR671846',
                           'SRR671847', 'SRR671848', 'SRR671849', 'SRR671850', 'SRR671851', 'SRR671852', 'SRR671853',
                           'SRR671854', 'SRR671855', 'SRR671856', 'SRR671857', 'SRR671858', 'SRR671859', 'SRR671860',
                           'SRR671861', 'SRR671862', 'SRR671863', 'SRR671864', 'SRR671865', 'SRR671866', 'SRR671867',
                           'SRR671868', 'SRR671869', 'SRR671870', 'SRR671871', 'SRR671872', 'SRR671873', 'SRR671874',
                           'SRR671875', 'SRR671876', 'SRR671877', 'SRR671878', 'SRR671879']},
    #
    {
        'Source': "Global expansion of Mycobacterium tuberculosis lineage 4 shaped by colonial migration and local adaptation",
        'Author': "Brynildsrud et al.",
        'study accession number': '',
        'run accessions': ['ERR760595', 'ERR760596', 'ERR760597', 'ERR760598', 'ERR760599', 'ERR760600', 'ERR760601',
                           'ERR760602', 'ERR757162', 'ERR757163', 'ERR757164', 'ERR760603', 'ERR760604', 'ERR757165',
                           'ERR760605', 'ERR757166', 'ERR757167', 'ERR850262', 'ERR760607', 'ERR760608', 'ERR760609',
                           'ERR760610', 'ERR760611', 'ERR760612', 'ERR760731', 'ERR757168', 'ERR760732', 'ERR757169',
                           'ERR760733', 'ERR757170', 'ERR757171', 'ERR757172', 'ERR760734', 'ERR757173', 'ERR757174',
                           'ERR757175', 'ERR760735', 'ERR757176', 'ERR757177', 'ERR760736', 'ERR757178', 'ERR757179',
                           'ERR757180', 'ERR757181', 'ERR757182', 'ERR760737', 'ERR760738', 'ERR757183', 'ERR757184',
                           'ERR850078', 'ERR757185', 'ERR760739', 'ERR757186', 'ERR760740', 'ERR760741', 'ERR760742',
                           'ERR760743', 'ERR760744', 'ERR757187', 'ERR760745', 'ERR760746', 'ERR760747', 'ERR757188',
                           'ERR760748', 'ERR760749', 'ERR760750', 'ERR760751', 'ERR760752', 'ERR760753', 'ERR760754',
                           'ERR757189', 'ERR760755', 'ERR760809', 'ERR760756', 'ERR760757', 'ERR760758', 'ERR760759',
                           'ERR760760', 'ERR760761', 'ERR760762', 'ERR760763', 'ERR760764', 'ERR760765', 'ERR760766',
                           'ERR760767', 'ERR760768', 'ERR845303', 'ERR845303', 'ERR845304', 'ERR845305', 'ERR845305',
                           'ERR845306', 'ERR760810', 'ERR845307', 'ERR845307', 'ERR845308', 'ERR845308', 'ERR845327',
                           'ERR845327', 'ERR760811', 'ERR760812', 'ERR760813', 'ERR760814', 'ERR845328', 'ERR845328',
                           'ERR845329', 'ERR845329', 'ERR762352', 'ERR760778', 'ERR760815', 'ERR760779', 'ERR760780',
                           'ERR760781', 'ERR760782', 'ERR760783', 'ERR760780', 'ERR760784', 'ERR760785', 'ERR760786',
                           'ERR760787', 'ERR760788', 'ERR760789', 'ERR760790', 'ERR760791', 'ERR760816', 'ERR760792',
                           'ERR760793', 'ERR760794', 'ERR760795', 'ERR760817', 'ERR760818', 'ERR760796', 'ERR757151',
                           'ERR760797', 'ERR760798', 'ERR762353', 'ERR762353', 'ERR760799', 'ERR760800', 'ERR760801',
                           'ERR760802', 'ERR760803', 'ERR760819', 'ERR760820', 'ERR760821', 'ERR760822', 'ERR760823',
                           'ERR760824', 'ERR757159', 'ERR760825', 'ERR757160', 'ERR760826', 'ERR760827', 'ERR760828',
                           'ERR757161', 'ERR760829', 'ERR760830', 'ERR760831', 'ERR760832', 'ERR760833', 'ERR760834',
                           'ERR760835', 'ERR760836', 'ERR760837', 'ERR760838', 'ERR760839', 'ERR760840', 'ERR760841',
                           'ERR760842', 'ERR760843', 'ERR760844', 'ERR760845', 'ERR757152', 'ERR760846', 'ERR760847',
                           'ERR760848', 'ERR760849', 'ERR760850', 'ERR760851', 'ERR757146', 'ERR760852', 'ERR760853',
                           'ERR760854', 'ERR760855', 'ERR760856', 'ERR760857', 'ERR760858', 'ERR760859', 'ERR760860',
                           'ERR760861', 'ERR760862', 'ERR760863', 'ERR760864', 'ERR760865', 'ERR760866', 'ERR760867',
                           'ERR760868', 'ERR760869', 'ERR760870', 'ERR760871', 'ERR760872', 'ERR760873', 'ERR760874',
                           'ERR760875', 'ERR760876', 'ERR760877', 'ERR760878', 'ERR760879', 'ERR760880', 'ERR760881',
                           'ERR760882', 'ERR760883', 'ERR760884', 'ERR760885', 'ERR760886', 'ERR760887', 'ERR760888',
                           'ERR760889', 'ERR760890', 'ERR760891', 'ERR760892', 'ERR760893', 'ERR760894', 'ERR760895',
                           'ERR760896', 'ERR760897', 'ERR845330', 'ERR760899', 'ERR760900', 'ERR760901', 'ERR760902',
                           'ERR760903', 'ERR760904', 'ERR760905', 'ERR760906', 'ERR760907', 'ERR760908', 'ERR760909',
                           'ERR760910', 'ERR760911', 'ERR760912', 'ERR760913', 'ERR760914', 'ERR760915', 'ERR760916',
                           'ERR760917', 'ERR760918', 'ERR760919', 'ERR760920', 'ERR760921', 'ERR760922', 'ERR760923',
                           'ERR760924', 'ERR760925', 'ERR757145', 'ERR757147', 'ERR757148', 'ERR757150', 'ERR760926',
                           'ERR760927', 'ERR776665', 'ERR772257', 'ERR772254', 'ERR776452', 'ERR776453', 'ERR776454',
                           'ERR772258', 'ERR776455', 'ERR776456', 'ERR776457', 'ERR776458', 'ERR776459', 'ERR776460',
                           'ERR776659', 'ERR776461', 'ERR776462', 'ERR776463', 'ERR776464', 'ERR772259', 'ERR776465',
                           'ERR776466', 'ERR776467', 'ERR776468', 'ERR776469', 'ERR776660', 'ERR776470', 'ERR776471',
                           'ERR776472', 'ERR776473', 'ERR772260', 'ERR776474', 'ERR772261', 'ERR772262', 'ERR776475',
                           'ERR776476', 'ERR772255', 'ERR776666', 'ERR776667', 'ERR776668', 'ERR776669', 'ERR776670',
                           'ERR776671', 'ERR772263', 'ERR776672', 'ERR776673', 'ERR776674', 'ERR776661', 'ERR772264',
                           'ERR776675', 'ERR772265', 'ERR776662', 'ERR776663', 'ERR772256', 'ERR776664', 'ERR2652988',
                           'ERR2652989', 'ERR2652990', 'ERR2652991', 'ERR2652992', 'ERR2652993', 'ERR2652994',
                           'ERR2652995', 'ERR2652996', 'ERR2652997', 'ERR2652998', 'ERR2652976', 'ERR2652999',
                           'ERR2653088', 'ERR2653087', 'ERR2653086', 'ERR2653085', 'ERR2653084', 'ERR2653083',
                           'ERR2652980', 'ERR2653082', 'ERR2653081', 'ERR2652977', 'ERR2653080', 'ERR2653079',
                           'ERR2653078', 'ERR2653077', 'ERR2653076', 'ERR2653089', 'ERR2653090', 'ERR2653091',
                           'ERR2652978', 'ERR2653092', 'ERR2653093', 'ERR2653094', 'ERR2653095', 'ERR2653096',
                           'ERR2653097', 'ERR2653098', 'ERR2653099', 'ERR2653100', 'ERR2652979', 'ERR2653101',
                           'ERR2653102', 'ERR2653103', 'ERR2653104', 'ERR2653105', 'ERR2653106', 'ERR2653107',
                           'ERR2653108', 'ERR2653000', 'ERR2653109', 'ERR2653110', 'ERR2653111', 'ERR2653112',
                           'ERR2653113', 'ERR2653114', 'ERR2653115', 'ERR2653116', 'ERR2653117', 'ERR2653118',
                           'ERR2653001', 'ERR2653168', 'ERR2653169', 'ERR2653170', 'ERR2653171', 'ERR2653172',
                           'ERR2653173', 'ERR2653174', 'ERR2653175', 'ERR2653002', 'ERR2653176', 'ERR2653177',
                           'ERR2653178', 'ERR2653179', 'ERR2653180', 'ERR2653181', 'ERR2653182', 'ERR2653003',
                           'ERR2653183', 'ERR2653184', 'ERR2653185', 'ERR2653186', 'ERR2653187', 'ERR2653188',
                           'ERR2653189', 'ERR2653190', 'ERR2653004', 'ERR2653191', 'ERR2653192', 'ERR2653193',
                           'ERR2653194', 'ERR2653195', 'ERR2653196', 'ERR2653197', 'ERR2653005', 'ERR2653198',
                           'ERR2653199', 'ERR2653200', 'ERR2653119', 'ERR2653120', 'ERR2653121', 'ERR2653122',
                           'ERR2653123', 'ERR2653124', 'ERR2653006', 'ERR2653125', 'ERR2653126', 'ERR2653127',
                           'ERR2653128', 'ERR2653129', 'ERR2653130', 'ERR2653131', 'ERR2653132', 'ERR2652981',
                           'ERR2653133', 'ERR2653134', 'ERR2653007', 'ERR2653135', 'ERR2653136', 'ERR2653137',
                           'ERR2653138', 'ERR2653139', 'ERR2653140', 'ERR2653141', 'ERR2653142', 'ERR2653143',
                           'ERR2653008', 'ERR2653144', 'ERR2653145', 'ERR2653146', 'ERR2653147', 'ERR2653148',
                           'ERR2653149', 'ERR2653150', 'ERR2653151', 'ERR2653152', 'ERR2653009', 'ERR2653153',
                           'ERR2653154', 'ERR2653155', 'ERR2653156', 'ERR2653157', 'ERR2653158', 'ERR2653159',
                           'ERR2653160', 'ERR2653161', 'ERR2653010', 'ERR2653162', 'ERR2653163', 'ERR2653164',
                           'ERR2653165', 'ERR2653166', 'ERR2653167', 'ERR2653201', 'ERR2653212', 'ERR2653223',
                           'ERR2653011', 'ERR2653233', 'ERR2653237', 'ERR2653238', 'ERR2653239', 'ERR2653240',
                           'ERR2653241', 'ERR2653202', 'ERR2653203', 'ERR2653204', 'ERR2653012', 'ERR2653205',
                           'ERR2653206', 'ERR2653207', 'ERR2653208', 'ERR2653209', 'ERR2653210', 'ERR2653211',
                           'ERR2653213', 'ERR2653214', 'ERR2653215', 'ERR2653013', 'ERR2653216', 'ERR2653217',
                           'ERR2653218', 'ERR2653219', 'ERR2653220', 'ERR2653221', 'ERR2653222', 'ERR2653224',
                           'ERR2653225', 'ERR2653014', 'ERR2653226', 'ERR2653227', 'ERR2653228', 'ERR2653229',
                           'ERR2653230', 'ERR2653231', 'ERR2653232', 'ERR2653234', 'ERR2653015', 'ERR2653235',
                           'ERR2653236', 'ERR2653016', 'ERR2652982', 'ERR2653017', 'ERR2653018', 'ERR2653019',
                           'ERR2653020', 'ERR2653021', 'ERR2653022', 'ERR2653023', 'ERR2653024', 'ERR2653025',
                           'ERR2652983', 'ERR2653026', 'ERR2653027', 'ERR2653028', 'ERR2653029', 'ERR2653030',
                           'ERR2653031', 'ERR2653032', 'ERR2653033', 'ERR2653034', 'ERR2652984', 'ERR2653035',
                           'ERR2653036', 'ERR2653037', 'ERR2653038', 'ERR2653039', 'ERR2653040', 'ERR2653041',
                           'ERR2653042', 'ERR2653043', 'ERR2652985', 'ERR2653044', 'ERR2653045', 'ERR2653046',
                           'ERR2653047', 'ERR2653048', 'ERR2653049', 'ERR2653050', 'ERR2653051', 'ERR2653052',
                           'ERR2653053', 'ERR2653054', 'ERR2653055', 'ERR2653056', 'ERR2653057', 'ERR2653058',
                           'ERR2653059', 'ERR2653060', 'ERR2652986', 'ERR2653061', 'ERR2653062', 'ERR2653063',
                           'ERR2653064', 'ERR2653065', 'ERR2653066', 'ERR2653067', 'ERR2653068', 'ERR2653069',
                           'ERR2652987', 'ERR2653070', 'ERR2653071', 'ERR2653072', 'ERR2653073', 'ERR2653074',
                           'ERR2653075', 'ERR2652926', 'ERR2652925', 'ERR2652927', 'ERR2652928', 'ERR2652929',
                           'ERR2652930', 'ERR2652931', 'ERR2652932', 'ERR2652933', 'ERR2652934', 'ERR2652935',
                           'ERR2652936', 'ERR2652937', 'ERR2652938', 'ERR2652939', 'ERR2652940', 'ERR2652941',
                           'ERR2652942', 'ERR2652943', 'ERR2652944', 'ERR2652945', 'ERR2652946', 'ERR2652947',
                           'ERR2652948', 'ERR2652949', 'ERR2652950', 'ERR2652951', 'ERR2652952', 'ERR2652953',
                           'ERR2652954', 'ERR2652955', 'ERR2652956', 'ERR2652957', 'ERR2652958', 'ERR2652959',
                           'ERR2652961', 'ERR2652960', 'ERR2652962', 'ERR2652963', 'ERR2652964', 'ERR2652965',
                           'ERR2652966', 'ERR2652967', 'ERR2652968', 'ERR2652969', 'ERR2652970', 'ERR2652971',
                           'ERR2652972', 'ERR2652973', 'ERR2652974', 'ERR2652975', 'ERR551789', 'ERR551791',
                           'ERR551583', 'ERR553103', 'ERR550723', 'ERR552392', 'ERR552393', 'ERR551929', 'ERR553057',
                           'ERR552529', 'ERR552617', 'ERR552618', 'ERR550798', 'ERR552192', 'ERR551590', 'ERR553229',
                           'ERR553151', 'ERR551502', 'ERR553306', 'ERR551820', 'ERR552919', 'ERR553317', 'ERR551676',
                           'ERR551677', 'ERR551092', 'ERR551354', 'ERR551073', 'ERR551416', 'ERR551961', 'ERR550819',
                           'ERR551950', 'ERR551253', 'ERR551587', 'ERR552492', 'ERR551075', 'ERR551076', 'ERR550775',
                           'ERR550622', 'ERR552592', 'ERR550939', 'ERR553370', 'ERR552643', 'ERR551367', 'ERR550697',
                           'ERR551039', 'ERR551597', 'ERR553380', 'ERR551808', 'ERR553004', 'ERR552448', 'ERR551685',
                           'ERR552634', 'ERR550879', 'ERR551687', 'ERR553104', 'ERR553087', 'ERR551359', 'ERR553316',
                           'ERR552869', 'ERR551404', 'ERR553041', 'ERR979062', 'ERR979063', 'ERR979065', 'ERR979066',
                           'ERR979067', 'ERR979068', 'ERR979069', 'ERR979070', 'ERR979071', 'ERR979072', 'ERR979073',
                           'ERR979074', 'ERR979075', 'ERR245649', 'ERR245650', 'ERR245653', 'ERR245658', 'ERR245659',
                           'ERR245661', 'ERR245664', 'ERR245665', 'ERR245666', 'ERR245667', 'ERR245673', 'ERR245681',
                           'ERR245683', 'ERR245685', 'ERR245686', 'ERR245690', 'ERR245693', 'ERR245697', 'ERR245700',
                           'ERR245701', 'ERR245706', 'ERR245707', 'ERR245708', 'ERR245709', 'ERR245711', 'ERR245712',
                           'ERR245714', 'ERR245718', 'ERR245719', 'ERR245724', 'ERR245725', 'ERR245729', 'ERR245730',
                           'ERR245731', 'ERR245733', 'ERR245736', 'ERR245737', 'ERR245738', 'ERR245740', 'ERR245741',
                           'ERR245743', 'ERR245744', 'ERR245745', 'ERR245746', 'ERR245749', 'ERR245750', 'ERR245752',
                           'ERR245755', 'ERR245756', 'ERR245758', 'ERR245759', 'ERR028608', 'ERR028628', 'ERR029204',
                           'ERR028620', 'ERR028621', 'ERR029201', 'ERR257886', 'ERR023747', 'ERR024355', 'ERR025431',
                           'ERR025450', 'ERR025453', 'ERR025458', 'ERR025438', 'ERR024345', 'ERR023759', 'ERR023764',
                           'ERR024340', 'ERR025422', 'ERR025423', 'ERR025433', 'ERR025457', 'ERR023737', 'ERR025428',
                           'ERR257909', 'ERR024343', 'ERR023738', 'ERR026472', 'ERR025441', 'ERR023731', 'ERR025424',
                           'ERR025415', 'ERR257880', 'ERR025454', 'ERR024344', 'ERR025460', 'ERR024346', 'ERR257891',
                           'ERR023757', 'ERR024347', 'ERR023760', 'ERR025434', 'ERR257892', 'ERR257910', 'ERR025435',
                           'ERR025459', 'ERR257893', 'ERR025451', 'ERR257894', 'ERR257895', 'ERR025436', 'ERR025427',
                           'ERR025455', 'ERR257931', 'ERR257932', 'ERR023761', 'ERR025440', 'ERR025432', 'ERR023739',
                           'ERR025420', 'ERR023740', 'ERR257896', 'ERR257897', 'ERR257898', 'ERR257911', 'ERR257912',
                           'ERR257899', 'ERR257915', 'ERR023732', 'ERR257900', 'ERR257901', 'ERR023758', 'ERR025421',
                           'ERR257902', 'ERR257903', 'ERR257904', 'ERR257917', 'ERR257933', 'ERR257934', 'ERR257905',
                           'ERR257906', 'ERR025414', 'ERR025417', 'ERR025418', 'ERR025419', 'ERR023729', 'ERR257907',
                           'ERR025452', 'ERR257935', 'ERR023746', 'ERR026475', 'ERR025426', 'ERR025456', 'ERR025416',
                           'ERR025429', 'ERR026476', 'ERR025430', 'ERR257908', 'ERR024359', 'ERR028610', 'ERR028613',
                           'ERR028612', 'ERR028615', 'ERR028616', 'ERR028617', 'ERR028622', 'ERR029211', 'ERR029202',
                           'ERR028624', 'ERR028626', 'ERR028625', 'ERR028629', 'ERR751351', 'ERR751354', 'ERR751355',
                           'ERR751357', 'ERR751359', 'ERR751362', 'ERR751363', 'ERR751365', 'ERR751367', 'ERR751368',
                           'ERR751372', 'ERR751373', 'ERR751374', 'ERR751375', 'ERR751377', 'ERR751382', 'ERR751384',
                           'ERR751387', 'ERR751389', 'ERR751394', 'ERR751399', 'ERR751400', 'ERR751401', 'ERR751402',
                           'ERR751403', 'ERR751406', 'ERR751409', 'ERR751415', 'ERR751416', 'ERR751417', 'ERR751418',
                           'ERR751423', 'ERR751424', 'ERR751425', 'ERR751426', 'ERR751427', 'ERR751429', 'ERR751431',
                           'ERR751434', 'ERR751435', 'ERR751436', 'ERR751437', 'ERR751439', 'ERR751442', 'ERR751444',
                           'ERR751445', 'ERR751446', 'ERR751448', 'ERR751449', 'ERR751450', 'ERR751455', 'ERR751456',
                           'ERR751457', 'ERR751458', 'ERR751459', 'ERR751460', 'ERR751462', 'ERR751463', 'ERR751468',
                           'ERR751469', 'ERR751472', 'ERR751474', 'ERR751476', 'ERR751477', 'ERR751480', 'ERR751482',
                           'ERR751485', 'ERR751487', 'ERR751489', 'ERR751492', 'ERR751494', 'ERR751496', 'ERR751498',
                           'ERR751499', 'ERR751500', 'ERR751501', 'ERR751504', 'ERR751507', 'ERR751508', 'ERR751509',
                           'ERR751511', 'ERR751512', 'ERR751513', 'ERR751515', 'ERR751523', 'ERR751524', 'ERR751528',
                           'ERR751530', 'ERR751533', 'ERR751535', 'ERR751536', 'ERR751538', 'ERR751541', 'ERR751542',
                           'ERR751543', 'ERR751544', 'ERR751545', 'ERR751547', 'ERR751548', 'ERR751549', 'ERR751550',
                           'ERR751551', 'ERR751553', 'ERR751554', 'ERR751556', 'ERR751557', 'ERR751559', 'ERR751560',
                           'ERR751561', 'ERR751562', 'ERR751563', 'ERR751564', 'ERR751565', 'ERR751566', 'ERR751567',
                           'ERR751568', 'ERR751569', 'ERR751570', 'ERR751571', 'ERR751572', 'ERR751573', 'ERR751574',
                           'ERR751578', 'ERR751580', 'ERR751581', 'ERR751582', 'ERR751585', 'ERR751586', 'ERR751587',
                           'ERR751589', 'ERR751592', 'ERR751593', 'ERR751595', 'ERR751596', 'ERR751597', 'ERR751598',
                           'ERR751599', 'ERR751600', 'ERR751604', 'ERR751605', 'ERR751606', 'ERR751607', 'ERR751608',
                           'ERR751609', 'ERR751612', 'ERR751614', 'ERR751616', 'ERR751617', 'ERR751619', 'ERR751621',
                           'ERR751622', 'ERR751623', 'ERR751624', 'ERR751626', 'ERR751628', 'ERR751629', 'ERR751630',
                           'ERR751631', 'ERR751633', 'ERR775293', 'ERR775294', 'ERR775296', 'ERR775299', 'ERR775300',
                           'ERR775302', 'ERR775305', 'ERR775311', 'ERR775315', 'ERR775316', 'ERR775317', 'ERR775318',
                           'ERR775326', 'ERR775328', 'ERR775329', 'ERR775330', 'ERR775331', 'ERR775332', 'ERR775336',
                           'ERR775337', 'ERR775338', 'ERR775340', 'ERR775342', 'ERR775345', 'ERR775348', 'ERR775350',
                           'ERR775351', 'ERR775353', 'ERR775354', 'ERR775357', 'ERR775358', 'ERR775360', 'ERR775361',
                           'ERR775365', 'ERR775367', 'ERR775369', 'ERR775371', 'ERR775372', 'ERR775374', 'ERR775383',
                           'ERR775387', 'ERR775388', 'ERR779842', 'ERR779844', 'ERR779847', 'ERR779848', 'ERR779850',
                           'ERR779854', 'ERR779855', 'ERR779858', 'ERR779860', 'ERR779861', 'ERR779864', 'ERR779865',
                           'ERR779867', 'ERR779868', 'ERR779869', 'ERR779870', 'ERR779871', 'ERR779873', 'ERR779874',
                           'ERR779875', 'ERR779877', 'ERR779878', 'ERR779880', 'ERR779884', 'ERR779885', 'ERR779886',
                           'ERR779887', 'ERR779890', 'ERR779892', 'ERR779895', 'ERR779897', 'ERR779898', 'ERR779900',
                           'ERR779901', 'ERR779902', 'ERR779906', 'ERR779907', 'ERR779908', 'ERR779909', 'ERR779911',
                           'ERR779912', 'ERR779915', 'ERR779917', 'ERR779921', 'ERR779922', 'ERR779925', 'ERR845937',
                           'ERR845941', 'ERR275181', 'ERR275182', 'ERR275183', 'ERR275184', 'ERR275185', 'ERR275186',
                           'ERR275187', 'ERR275189', 'ERR275190', 'ERR275191', 'ERR275192', 'ERR275193', 'ERR275195',
                           'ERR275196', 'ERR275197', 'ERR275198', 'ERR275199', 'ERR275200', 'ERR275201', 'ERR275202',
                           'ERR275203', 'ERR275204', 'ERR275206', 'ERR275207', 'ERR275208', 'ERR275209', 'ERR275210',
                           'ERR275211', 'ERR275212', 'ERR275213', 'ERR275214', 'ERR275217', 'ERR275218', 'ERR275219',
                           'ERR275220', 'ERR275221', 'ERR275222', 'ERR275223', 'ERR275224', 'ERR275225', 'ERR275227',
                           'ERR275228', 'ERR275229', 'ERR275230', 'ERR275231', 'ERR275232', 'ERR275233', 'ERR275234',
                           'ERR275235', 'ERR275236', 'SRR1186316', 'SRR1186993', 'SRR1187001', 'SRR1187011',
                           'SRR1187028', 'SRR1187037', 'SRR1187076', 'SRR1187082', 'SRR1187085', 'SRR1187086',
                           'SRR1187087', 'SRR1187088', 'SRR1187089', 'SRR1187181', 'SRR1187183', 'SRR1187184',
                           'SRR1187186', 'SRR1187192', 'SRR1187195', 'SRR1187196', 'SRR1187251', 'SRR1187252',
                           'SRR1187295', 'SRR1187297', 'SRR1187380', 'SRR1187393', 'SRR1187423', 'SRR1187442',
                           'SRR1187576', 'SRR1187598', 'SRR1187618', 'SRR1187619', 'SRR1187620', 'SRR1187627',
                           'SRR1187628', 'SRR1187629', 'SRR1187630', 'SRR1187631', 'SRR1187633', 'SRR1187945',
                           'SRR1187946', 'SRR1187983', 'SRR1188082', 'SRR1188083', 'SRR1188084', 'SRR1188085',
                           'SRR1188086', 'SRR1188087', 'SRR1188119', 'SRR1188121', 'SRR1188127', 'SRR1188130',
                           'SRR1188131', 'SRR1188133', 'SRR1188137', 'SRR1188138', 'SRR1188143', 'SRR1188170',
                           'SRR1188174', 'SRR1188175', 'SRR1188181', 'SRR1188183', 'SRR1188184', 'SRR1188188',
                           'SRR1188220', 'SRR1188259', 'SRR1188284', 'SRR1188286', 'SRR1188339', 'SRR1188341',
                           'SRR1188343', 'SRR1188359', 'SRR1188439', 'SRR1188459', 'SRR1188479', 'SRR1188486',
                           'SRR1188487', 'SRR1188492', 'SRR1188530', 'SRR1190432', 'SRR1190434', 'SRR1190442',
                           'SRR1190468', 'SRR1190470', 'SRR1190473', 'SRR1190474', 'SRR1190475', 'SRR1190477',
                           'SRR1190478', 'SRR1190480', 'SRR1191277', 'SRR1191298', 'SRR1191488', 'SRR1191489',
                           'SRR1191546', 'SRR1191547', 'SRR1191666', 'SRR1191727', 'SRR1200251', 'SRR1367198',
                           'SRR1367199', 'SRR1367200', 'SRR1640242', 'SRR1640246', 'SRR1640250', 'SRR1640253',
                           'SRR1640254', 'SRR1640266', 'SRR1640273', 'SRR1640287', 'SRR1640289', 'SRR1640290',
                           'SRR1640291', 'SRR1640292', 'SRR1640293', 'SRR1640294', 'SRR1640295', 'SRR1640316',
                           'SRR1640321', 'SRR1640326', 'SRR1640327', 'SRR1640334', 'SRR1640335', 'SRR1640336',
                           'SRR1640338', 'SRR1640340', 'SRR1640342', 'SRR1640343', 'SRR1640345', 'SRR1640346',
                           'SRR1640347', 'SRR1640348', 'SRR1640349', 'SRR1640539', 'SRR1640540', 'SRR1640541',
                           'SRR1640542', 'SRR1640543', 'SRR1640547', 'SRR1640549', 'SRR1640551', 'SRR1640562',
                           'SRR1640563', 'SRR1640564', 'SRR1640565', 'SRR1640566', 'SRR1640567', 'SRR1640569',
                           'SRR1640570', 'SRR1640571', 'SRR1640574', 'SRR1640696', 'SRR1640697', 'SRR1640701',
                           'SRR1640706', 'SRR1640707', 'SRR1640708', 'SRR1640709', 'SRR1640710', 'SRR1640711',
                           'SRR1640712', 'SRR1640713', 'SRR1640714', 'SRR1640749', 'SRR1640751', 'SRR1640752',
                           'ERR067614', 'ERR067615', 'ERR067616', 'ERR067617', 'ERR067601', 'ERR067602', 'ERR067767',
                           'ERR067719', 'ERR067577', 'ERR067579', 'ERR067585', 'ERR067607', 'ERR067603', 'ERR067609',
                           'ERR067606', 'ERR067612', 'ERR067693', 'ERR067697', 'ERR067735', 'ERR067737', 'ERR067624',
                           'ERR067747', 'ERR067725', 'ERR067726', 'ERR067727', 'ERR067629', 'ERR067748', 'ERR067751',
                           'ERR067591', 'ERR067752', 'ERR067753', 'ERR067756', 'ERR067757', 'ERR067728', 'ERR067729',
                           'ERR067730', 'ERR067731', 'ERR067732', 'ERR067763', 'ERR067733', 'ERR067734', 'ERR067596',
                           'ERR067597', 'ERR067599', 'ERR017793', 'ERR017800', 'ERR027469', 'ERR067657', 'ERR067658',
                           'ERR067661', 'ERR067664', 'ERR067666', 'ERR067636', 'ERR067670', 'ERR067640', 'ERR067645',
                           'ERR067675', 'ERR067632', 'ERR067634', 'ERR067676', 'ERR067677', 'ERR067678', 'ERR067679',
                           'ERR067700', 'ERR067703', 'ERR067706', 'ERR067708', 'ERR067716', 'ERR067680', 'ERR067682',
                           'ERR067683', 'ERR067684', 'ERR067685', 'ERR067690', 'ERR067691', 'SRR1047970', 'SRR1047971',
                           'SRR1047972', 'SRR833028', 'SRR833121', 'SRR833024', 'SRR832991', 'SRR833109', 'SRR924700',
                           'SRR924701', 'SRR924707', 'SRR1011465', 'SRR1180299', 'SRR1140950', 'ERR040113', 'ERR040116',
                           'ERR040117', 'ERR040128', 'ERR040130', 'ERR040142', 'ERR040125', 'ERR038740', 'ERR040120',
                           'ERR040122', 'ERR040124', 'ERR040127', 'ERR038739', 'ERR040131', 'ERR040132', 'ERR040134',
                           'ERR040135', 'ERR040138', 'ERR038744', 'ERR038736', 'ERR038737', 'ERR038738', 'ERR038741',
                           'ERR038742', 'ERR038745', 'ERR038747', 'ERR038748', 'ERR038749', 'ERR038750', 'ERR038754',
                           'ERR017794', 'ERR027464', 'ERR017798', 'ERR017797', 'ERR019871', 'ERR047885', 'ERR047886',
                           'ERR047881', 'ERR2652914', 'ERR2652915', 'ERR2652916', 'ERR2652917', 'ERR2652918',
                           'ERR2652919', 'ERR2652920', 'ERR2652921', 'ERR2652922', 'ERR2652923', 'ERR2652924',
                           'SRR5065225', 'SRR5065416', 'SRR5065335', 'SRR5065239', 'SRR5065316', 'SRR5065690',
                           'SRR5065701', 'SRR5065252', 'SRR5065396', 'SRR5065659', 'SRR5065476', 'SRR5065417',
                           'SRR5065263', 'SRR5065329', 'SRR5065248', 'SRR5065206', 'SRR5065539', 'SRR5065214',
                           'SRR5065577', 'SRR5065517', 'SRR5065578', 'SRR5065246', 'SRR5065571', 'SRR5065374',
                           'SRR5065602', 'SRR5065408', 'SRR5065267', 'SRR5065508', 'SRR5065649', 'SRR5065276',
                           'SRR5065363', 'SRR5065254', 'SRR5065546', 'SRR5065463', 'SRR5065639', 'SRR5065677',
                           'SRR5065242', 'SRR5065490', 'SRR5065399', 'SRR5065244', 'SRR5065455', 'SRR5065622',
                           'SRR5065279', 'SRR5065629', 'SRR5065453', 'SRR5065454', 'SRR5065493', 'SRR5065268',
                           'SRR5065233', 'SRR5065636', 'SRR5065323', 'SRR5065657', 'SRR5065492', 'SRR5065597',
                           'SRR5065542', 'SRR5065540', 'SRR5065295', 'SRR5065203', 'SRR5065230', 'SRR5065698',
                           'SRR5065223', 'SRR5067382', 'SRR5067284', 'SRR5067463', 'SRR5067225', 'SRR5067685',
                           'SRR5067274', 'SRR5067602', 'SRR5067634', 'SRR5067508', 'SRR5067377', 'SRR5067671',
                           'SRR5067625', 'SRR5067680', 'SRR5067621', 'SRR5067559', 'SRR5067224', 'SRR5067473',
                           'SRR5067407', 'SRR5067355', 'SRR5067358', 'SRR5067287', 'SRR5067652', 'SRR5067515',
                           'SRR5067303', 'SRR5067578', 'SRR5067575', 'SRR5067398', 'SRR5067361', 'SRR5067363',
                           'SRR5067435', 'SRR5067395', 'SRR5067384', 'SRR5067281', 'SRR5067577', 'SRR5067674',
                           'SRR5067622', 'SRR5067263', 'SRR5067389', 'SRR5067507', 'SRR5067656', 'SRR5067459',
                           'SRR5067534', 'SRR5067589', 'SRR5067418', 'SRR5067267', 'SRR5067392', 'SRR5067615',
                           'SRR5067497', 'SRR5067569', 'SRR5067415', 'SRR5067427', 'SRR5067394', 'SRR5067604',
                           'SRR5067373', 'SRR5067709', 'SRR5073561', 'SRR5073580', 'SRR5073785', 'SRR5073667',
                           'SRR5073565', 'SRR5073896', 'SRR5073887', 'SRR5073871', 'SRR5073733', 'SRR5073502',
                           'SRR5073729', 'SRR5073715', 'SRR5073877', 'SRR5073572', 'SRR5073979', 'SRR5073640',
                           'SRR5073566', 'SRR5073908', 'SRR5073855', 'SRR5073876', 'SRR5073781', 'SRR5073989',
                           'SRR5073620', 'SRR5073934', 'SRR5073957', 'SRR5073693', 'SRR5073947', 'SRR5073609',
                           'SRR5073811', 'SRR5073788', 'SRR5073837', 'SRR5073784', 'SRR5073746', 'SRR5073553',
                           'SRR5073825', 'SRR5073894', 'SRR5073700', 'SRR5073897', 'SRR5073960', 'SRR5073765',
                           'SRR5073827', 'SRR5073817', 'SRR5073842', 'SRR5073556', 'SRR5073978', 'SRR5073981',
                           'SRR5073910', 'SRR5073529', 'SRR5073624', 'SRR5073688', 'SRR5073804', 'SRR5073991',
                           'SRR5073966', 'SRR5073805', 'SRR5073950', 'SRR5073516', 'SRR5073711', 'SRR5073999',
                           'SRR5073770', 'SRR5073940', 'SRR5073627', 'SRR5073675', 'SRR5073623', 'SRR5074192',
                           'SRR5074109', 'SRR5074178', 'SRR5074072', 'SRR5074190', 'SRR5074142', 'SRR5074075',
                           'SRR5074165', 'SRR5074111', 'SRR5074187', 'SRR5074141', 'SRR5074102', 'ERR650569',
                           'ERR651000', 'ERR651004', 'SRR1188186']},
    #
    {'Source': "Genetic diversity of Lineage 7 isolates in Ethiopia",
     'Author': "University of Oslo",
     'study accession number': 'PRJEB13960',
     'location': 'Ethiopia',
     'run accessions': ['ERR1971849', 'ERR1971850', 'ERR1971852', 'ERR1971853', 'ERR1971854', 'ERR1971855',
                        'ERR1971856', 'ERR1971857', 'ERR1971858', 'ERR1971859', 'ERR1971860', 'ERR1971861',
                        'ERR1971862', 'ERR1971863', 'ERR1971864', 'ERR1971865', 'ERR1971866', 'ERR1971867',
                        'ERR1971868', 'ERR1971869', 'ERR1971870', 'ERR1971871', 'ERR1971872', 'ERR1971873',
                        'ERR1971874', 'ERR1971876', 'ERR1971877']},
    #
    {'Source': "Tuberculosis isolates from patients in Ethiopia that represents a new lineage (Lineage 7)",
     'Author': "CSISP",
     'study accession number': 'PRJEB3124',
     'location': 'Ethiopia',
     'run accessions': ['ERR159958', 'ERR159959']},
    #
    {'Source': "Norwegian-African large Lineage 3 M. tuberculosis cluster",
     'Author': "Norwegian institute of public health",
     'study accession number': 'PRJEB23495',
     'run accessions': ['ERR2245275', 'ERR2245276', 'ERR2245277', 'ERR2245278', 'ERR2245279', 'ERR2245280',
                        'ERR2245281', 'ERR2245282', 'ERR2245283', 'ERR2245284', 'ERR2245285', 'ERR2245286',
                        'ERR2245287', 'ERR2245288', 'ERR2245289', 'ERR2245290', 'ERR2245291', 'ERR2245292',
                        'ERR2245293', 'ERR2245294', 'ERR2245295', 'ERR2245296', 'ERR2245297', 'ERR2245298',
                        'ERR2245299', 'ERR2245300', 'ERR2245301', 'ERR2245302', 'ERR2245303', 'ERR2245304',
                        'ERR2245305', 'ERR2245306', 'ERR2245307', 'ERR2245332', 'ERR2245333', 'ERR2245334',
                        'ERR2245335', 'ERR2245336', 'ERR2245337', 'ERR2245338', 'ERR2245339', 'ERR2245340',
                        'ERR2245341', 'ERR2245342', 'ERR2245343', 'ERR2245344', 'ERR2245345', 'ERR2245346',
                        'ERR2245347', 'ERR2245348', 'ERR2245349', 'ERR2245350', 'ERR2245351', 'ERR2245352',
                        'ERR2245353', 'ERR2245354', 'ERR2245355', 'ERR2245356', 'ERR2245357', 'ERR2245358',
                        'ERR2245359', 'ERR2245360', 'ERR2245361', 'ERR2245362', 'ERR2245363', 'ERR2245364',
                        'ERR2245365', 'ERR2245366', 'ERR2245367', 'ERR2245372', 'ERR2245373', 'ERR2245374',
                        'ERR2245375', 'ERR2245376', 'ERR2245377', 'ERR2245378', 'ERR2245379', 'ERR2245380',
                        'ERR2245381', 'ERR2245382', 'ERR2245383', 'ERR2245384', 'ERR2245385', 'ERR2245386',
                        'ERR2245387', 'ERR2245388', 'ERR2245389', 'ERR2245390', 'ERR2245391', 'ERR2245392',
                        'ERR2245393', 'ERR2245394', 'ERR2245395', 'ERR2245396', 'ERR2245397', 'ERR2245398',
                        'ERR2245399', 'ERR2245400', 'ERR2245401', 'ERR2245402', 'ERR2245403', 'ERR2245404',
                        'ERR2245405', 'ERR2245406', 'ERR2245407', 'ERR2245408', 'ERR2245409', 'ERR2245410',
                        'ERR2245411', 'ERR2245412', 'ERR2245413', 'ERR2245414', 'ERR2245415', 'ERR2245416',
                        'ERR2245417', 'ERR2245418', 'ERR2245419', 'ERR2245420', 'ERR2245421', 'ERR2245422',
                        'ERR2245423', 'ERR2245424', 'ERR2245425', 'ERR2245426', 'ERR2245427', 'ERR2512375',
                        'ERR2512376', 'ERR2512377', 'ERR2512378']},
    #
    {'Source': "Sequencing of ancient DNA of Mycobacterium tuberculosis from archaeological specimens",
     'Author': "UOM",
     'study accession number': 'PRJEB3329',
     'run accessions': ['ERR179757']},
    #
    {
        'Source': "Multidrug and extensively drug resistant tuberculosis among refugees, migrant workers and residents in Lebanon",
        'Author': "El Achkar S, Demanche C, Osman M, Rafei R, Ismail MB, Yaacoub H, Pinçon C, Duthoy S, De Matos F, Gaudin C, Trovato A, Cirillo DM, Hamze M, Supply P.",
        'study accession number': 'PRJNA488372',
        'location': 'Lebanon',
        'run accessions': ['SRR7765062', 'SRR7765063']},
    #
    {
        'Source': "Reference set of Mycobacterium tuberculosis clinical strains: A tool for research and product development",
        'Author': "Sònia Borrell, Andrej Trauner, Daniela Brites, Leen Rigouts, Chloe Loiseau, Mireia Coscolla, Stefan Niemann, Bouke De Jong, Dorothy Yeboah-Manu, Midori Kato-Maeda, Julia Feldmann, Miriam Reinhard, Christian Beisel, Sebastien Gagneux",
        'study accession number': '',
        'run accessions': ['ERR2704679', 'ERR2704680', 'ERR2704704', 'ERR2704676', 'ERR2704677', 'ERR2704685',
                           'ERR2704699', 'ERR2704698', 'ERR2704702', 'ERR2704701', 'ERR2704683', 'ERR2704703',
                           'ERR2704684', 'ERR2704675', 'ERR2704696', 'ERR2704697', 'ERR2704678', 'ERR2704693',
                           'ERR2704682', 'ERR2704686', 'ERR2704706', 'ERR2704700', 'ERR2704705', 'ERR2704689',
                           'ERR2704709', 'ERR2704694', 'ERR2704690', 'ERR2704708', 'ERR2704707', 'ERR2704692',
                           'ERR2704691', 'ERR2704681', 'ERR2704687', 'ERR2704688', 'ERR2704711', 'ERR2704695',
                           'ERR2704710']},
    #
    {'Source': "Harmonized Genome Wide Typing of Tubercle Bacilli Using a Web-Based Gene-By-Gene Nomenclature System",
     'Author': "Thomas A. Kohl, Dag Harmsen, Jörg Rothgänger, Timothy Walker, Roland Diel, Stefan Niemann",
     'study accession number': '',
     'run accessions': ['ERR233352', 'ERR233358', 'ERR233379', 'ERR234170', 'ERR234222', 'ERR234259', 'ERR550640',
                        'ERR551491', 'ERR551666', 'SRR017677', 'ERR553084', 'ERR552680', 'ERR553271', 'ERR552212',
                        'ERR551655', 'ERR498366', 'ERR552202', 'ERR498356']},
    #
    {
        'Source': "Whole genome sequencing of drug resistant Mycobacterium tuberculosis isolates from a high burden tuberculosis region of North West Pakistan",
        'Author': "Abdul Jabbar, Jody E. Phelan, Paola Florez de Sessions, Taj Ali Khan, Hazir Rahman, Sadiq Noor Khan, Daire M. Cantillon, Leticia Muraro Wildner, Sajid Ali, Susana Campino, Simon J. Waddell 6, Taane G. Clark",
        'study accession number': '',
        'location': 'Pakistan',
        'run accessions': ['ERR3335723', 'ERR3335724', 'ERR3335725', 'ERR3335726', 'ERR3335727', 'ERR3335728',
                           'ERR3335729', 'ERR3335730', 'ERR3335731', 'ERR3335732', 'ERR3335733', 'ERR3335734',
                           'ERR3335735', 'ERR3335736', 'ERR3335737', 'ERR3335738', 'ERR3335739', 'ERR3335740',
                           'ERR3335741', 'ERR3335742', 'ERR3335743', 'ERR3335744', 'ERR3335745', 'ERR3335746',
                           'ERR3335747', 'ERR3335748', 'ERR3335749', 'ERR3335750', 'ERR3335751', 'ERR3335752',
                           'ERR3335753', 'ERR3335754', 'ERR3335755', 'ERR3335756', 'ERR3335757', 'ERR3335758',
                           'ERR3335759', 'ERR3335760', 'ERR3335761', 'ERR3335762', 'ERR3335763', 'ERR3335764',
                           'ERR3335765', 'ERR3335766', 'ERR3335767', 'ERR3335768', 'ERR3335769', 'ERR3335770']},
    #
    {
        'Source': "Eighteenth-century genomes show that mixed infections were common at time of peak tuberculosis in Europe",
        'Author': "Gemma L. Kay, Martin J. Sergeant, Zhemin Zhou, Jacqueline Z.-M. Chan, Andrew Millard, Joshua Quick, Ildikó Szikossy, Ildikó Pap, Mark Spigelman, Nicholas J. Loman, Mark Achtman, Helen D. Donoghue, and Mark J. Pallen",
        'study accession number': 'PRJEB7454',
        'run accessions': ['ERR650569', 'ERR650970', 'ERR650971', 'ERR650973', 'ERR650974', 'ERR650975', 'ERR650976',
                           'ERR650977', 'ERR650978', 'ERR650979', 'ERR650980', 'ERR650981', 'ERR650982', 'ERR650983',
                           'ERR650984', 'ERR650985', 'ERR650986', 'ERR650987', 'ERR650988', 'ERR650990', 'ERR650991',
                           'ERR650992', 'ERR650993', 'ERR650994', 'ERR650995', 'ERR650996', 'ERR650997', 'ERR650998',
                           'ERR650999', 'ERR651000', 'ERR651001', 'ERR651002', 'ERR651003', 'ERR651005', 'ERR651006',
                           'ERR651007', 'ERR651008', 'ERR651009', 'ERR651010', 'ERR650989', 'ERR651004', 'ERR650972']},
    #
    {'Source': "Mycobacterium tuberculosis isolates from Kinshasa,DRC",
     'Author': "",
     'study accession number': 'PRJEB27847',
     'location': 'Congo',
     'run accessions': ['ERR2706914', 'ERR2706930', 'ERR2706945', 'ERR2706978', 'ERR2707002', 'ERR2707015',
                        'ERR2707066', 'ERR2707097', 'ERR2707138', 'ERR2707170', 'ERR2706932', 'ERR2706962',
                        'ERR2706927', 'ERR2706931', 'ERR2706954', 'ERR2706955', 'ERR2706913', 'ERR2707228',
                        'ERR2707056', 'ERR2707133', 'ERR2706959', 'ERR2707084', 'ERR2707043', 'ERR2707081',
                        'ERR2707102', 'ERR2707001']},
    #
    {
        'Source': "Spoligotyping and whole-genome sequencing analysis of lineage 1 strains of Mycobacterium tuberculosis in Da Nang, Vietnam",
        'Author': "Minako Hijikata, Naoto Keicho, Le Van Duc, Shinji Maeda, Nguyen Thi Le Hang, Ikumi Matsushita, Seiya Kato",
        'study accession number': 'PRJDB6149',
        'location': 'Vietnam',
        'run accessions': ['DRR099684', 'DRR099686', 'DRR099689', 'DRR099692', 'DRR099683', 'DRR099685', 'DRR099687',
                           'DRR099688', 'DRR099690', 'DRR099691', 'DRR099693', 'DRR099694']},
    #
    {'Source': "Genomic analyses of the ancestral Manila family of Mycobacterium tuberculosis",
     'Author': "Xuehua Wan, Kent Koster, Lishi Qian, Edward Desmond, Richard Brostrom, Shaobin Hou, James T. Douglas",
     'study accession number': 'PRJNA254678',
     'location': 'Manilla',
     'run accessions': ['SRR1510036', 'SRR1510037', 'SRR1510038', 'SRR1510039', 'SRR1510040', 'SRR1510041',
                        'SRR1510042', 'SRR1510043', 'SRR1510044', 'SRR1510045', 'SRR1510046', 'SRR1510047',
                        'SRR1510048', 'SRR1510049', 'SRR1510050', 'SRR1510051', 'SRR1510052', 'SRR1510053',
                        'SRR1510054', 'SRR1510055', 'SRR1510056', 'SRR1510057', 'SRR1510058', 'SRR1510059',
                        'SRR1510060', 'SRR1510061', 'SRR1510062', 'SRR1510063', 'SRR1510064', 'SRR1510065',
                        'SRR1510066', 'SRR1510067', 'SRR1510068', 'SRR1510069', 'SRR1510070', 'SRR1510071',
                        'SRR1510072']},
    #
    {
        'Source': "Large-scale whole genome sequencing of M. tuberculosis provides insights into transmission in a high prevalence area.",
        'Author': "Guerra-Assunção JA et al.",
        'study accession number': 'PRJEB2358',
        'location': 'Malawi',
        'run accessions': ['ERR037468', 'ERR037473', 'ERR037476', 'ERR037477', 'ERR037483', 'ERR037484', 'ERR037485',
                           'ERR037486', 'ERR037487', 'ERR037488', 'ERR037489', 'ERR037490', 'ERR037492', 'ERR037493',
                           'ERR037494', 'ERR037495', 'ERR037496', 'ERR037497', 'ERR037498', 'ERR037499', 'ERR037500',
                           'ERR037501', 'ERR037502', 'ERR037503', 'ERR037504', 'ERR037505', 'ERR037506', 'ERR037507',
                           'ERR037508', 'ERR037509', 'ERR037510', 'ERR037511', 'ERR037512', 'ERR037513', 'ERR037514',
                           'ERR037515', 'ERR037516', 'ERR037517', 'ERR037518', 'ERR037519', 'ERR037520', 'ERR037521',
                           'ERR037522', 'ERR037523', 'ERR037524', 'ERR037525', 'ERR037527', 'ERR037528', 'ERR037529',
                           'ERR037530', 'ERR037531', 'ERR037532', 'ERR037533', 'ERR037534', 'ERR037535', 'ERR037536',
                           'ERR037537', 'ERR037538', 'ERR037539', 'ERR037540', 'ERR037541', 'ERR037542', 'ERR037543',
                           'ERR037544', 'ERR037545', 'ERR037546', 'ERR037547', 'ERR037548', 'ERR037549', 'ERR037550',
                           'ERR037551', 'ERR037552', 'ERR037553', 'ERR037554', 'ERR037555', 'ERR245646', 'ERR245647',
                           'ERR245648', 'ERR245649', 'ERR245650', 'ERR245651', 'ERR245652', 'ERR245653', 'ERR245654',
                           'ERR245655', 'ERR245657', 'ERR245658', 'ERR245659', 'ERR245660', 'ERR245661', 'ERR245662',
                           'ERR245663', 'ERR245665', 'ERR245666', 'ERR245667', 'ERR245668', 'ERR245669', 'ERR245670',
                           'ERR245671', 'ERR245672', 'ERR245673', 'ERR245674', 'ERR245675', 'ERR245676', 'ERR245677',
                           'ERR245678', 'ERR245679', 'ERR245680', 'ERR245681', 'ERR245683', 'ERR245684', 'ERR245685',
                           'ERR245686', 'ERR245687', 'ERR245688', 'ERR245689', 'ERR245690', 'ERR245691', 'ERR245692',
                           'ERR245693', 'ERR245694', 'ERR245695', 'ERR245697', 'ERR245698', 'ERR245699', 'ERR245700',
                           'ERR245701', 'ERR245702', 'ERR245703', 'ERR245704', 'ERR245705', 'ERR245706', 'ERR245707',
                           'ERR245708', 'ERR245709', 'ERR245710', 'ERR245711', 'ERR245712', 'ERR245713', 'ERR245714',
                           'ERR245715', 'ERR245716', 'ERR245717', 'ERR245718', 'ERR245719', 'ERR245720', 'ERR245721',
                           'ERR245722', 'ERR245723', 'ERR245724', 'ERR245725', 'ERR245728', 'ERR245729', 'ERR245730',
                           'ERR245731', 'ERR245732', 'ERR245733', 'ERR245734', 'ERR245735', 'ERR245736', 'ERR245737',
                           'ERR245738', 'ERR245739', 'ERR245740', 'ERR245741', 'ERR245742', 'ERR245743', 'ERR245744',
                           'ERR245745', 'ERR245746', 'ERR245747', 'ERR245748', 'ERR245749', 'ERR245750', 'ERR245751',
                           'ERR245752', 'ERR245753', 'ERR245754', 'ERR245755', 'ERR245756', 'ERR245757', 'ERR245758',
                           'ERR245759', 'ERR245761', 'ERR245762', 'ERR245763', 'ERR245764', 'ERR245765', 'ERR245766',
                           'ERR245767', 'ERR245768', 'ERR245769', 'ERR245770', 'ERR245772', 'ERR245773', 'ERR245774',
                           'ERR245775', 'ERR245776', 'ERR245777', 'ERR245778', 'ERR245780', 'ERR245781', 'ERR245782',
                           'ERR245783', 'ERR245784', 'ERR245785', 'ERR245786', 'ERR245787', 'ERR245788', 'ERR245789',
                           'ERR245790', 'ERR245791', 'ERR245793', 'ERR245794', 'ERR245795', 'ERR245796', 'ERR245798',
                           'ERR245799', 'ERR245800', 'ERR245801', 'ERR245802', 'ERR245803', 'ERR245804', 'ERR245805',
                           'ERR245806', 'ERR245807', 'ERR245809', 'ERR245810', 'ERR245811', 'ERR245812', 'ERR245813',
                           'ERR245814', 'ERR245815', 'ERR245816', 'ERR245817', 'ERR245818', 'ERR245819', 'ERR245820',
                           'ERR245821', 'ERR245823', 'ERR245824', 'ERR245825', 'ERR245826', 'ERR245827', 'ERR245828',
                           'ERR245829', 'ERR245830', 'ERR245831', 'ERR245832', 'ERR245833', 'ERR245834', 'ERR245835',
                           'ERR245836', 'ERR245837', 'ERR245838', 'ERR245839', 'ERR245840', 'ERR245841', 'ERR245842',
                           'ERR245843', 'ERR245844', 'ERR245845', 'ERR245846', 'ERR245847', 'ERR245848', 'ERR245849',
                           'ERR036186', 'ERR036187', 'ERR036189', 'ERR036190', 'ERR036191', 'ERR036192', 'ERR036193',
                           'ERR036194', 'ERR036195', 'ERR036196', 'ERR036197', 'ERR036198', 'ERR036199', 'ERR036200',
                           'ERR036202', 'ERR036203', 'ERR036204', 'ERR036205', 'ERR036206', 'ERR036207', 'ERR036208',
                           'ERR036209', 'ERR036210', 'ERR036211', 'ERR036212', 'ERR036213', 'ERR036214', 'ERR036215',
                           'ERR036217', 'ERR036218', 'ERR036219', 'ERR036220', 'ERR036221', 'ERR036222', 'ERR036223',
                           'ERR036224', 'ERR036225', 'ERR036226', 'ERR036227', 'ERR036228', 'ERR036229', 'ERR036230',
                           'ERR036231', 'ERR036232', 'ERR036233', 'ERR036235', 'ERR036236', 'ERR036237', 'ERR036238',
                           'ERR036239', 'ERR036240', 'ERR036241', 'ERR036242', 'ERR036243', 'ERR036244', 'ERR036245',
                           'ERR036247', 'ERR036248', 'ERR036249', 'ERR037467', 'ERR037469', 'ERR037470', 'ERR037471',
                           'ERR037472', 'ERR037474', 'ERR037475', 'ERR037479', 'ERR037480', 'ERR037481', 'ERR245771',
                           'ERR036185', 'ERR037478', 'ERR245656', 'ERR036201', 'ERR245822', 'ERR036188', 'ERR245696',
                           'ERR245797', 'ERR245664', 'ERR245760', 'ERR037526', 'ERR036234', 'ERR245726', 'ERR037491',
                           'ERR037482', 'ERR245682', 'ERR036246', 'ERR245792', 'ERR245808', 'ERR036216', 'ERR245727',
                           'ERR245779']},
    #
    {'Source': "Whole-genome sequencing of multidrug-resistant Mycobacterium tuberculosis from Myanmar",
     'Author': "Htin Lin Aung, Thanda Tun, Danesh Moradigaravand, Claudio U. Köser, Wint Wint Nyunt, Si Thu Aung, Thandar Lwin, Kyi Kyi Thinn, John A. Crump, Julian Parkhill, Sharon J. Peacock, Gregory M. Cook, and Philip C. Hill",
     'study accession number': 'PRJEB10037',
     'location': 'Burma',
     'run accessions': ['ERR983228', 'ERR983229', 'ERR983230', 'ERR983231', 'ERR983232', 'ERR983233', 'ERR983234',
                        'ERR983235', 'ERR983236', 'ERR983237', 'ERR983238', 'ERR983239', 'ERR983240', 'ERR983241']},
    #
    {
        'Source': "Whole Genome Sequencing Based Characterization of Extensively Drug-Resistant Mycobacterium tuberculosis Isolates from Pakistan",
        'Author': "Asho Ali, Zahra Hasan, Ruth McNerney, Kim Mallard, Grant Hill-Cawthorne, Francesc Coll, Mridul Nair, Arnab Pain, Taane G. Clark, Rumina Hasan",
        'study accession number': 'PRJEB7798',
        'location': 'Pakistan',
        'run accessions': ['ERR688008', 'ERR688009', 'ERR688010', 'ERR688011', 'ERR688012', 'ERR688013', 'ERR688014',
                           'ERR688015', 'ERR688016', 'ERR688017', 'ERR688018', 'ERR688019', 'ERR688020', 'ERR688021',
                           'ERR688022', 'ERR688023', 'ERR688024', 'ERR688025', 'ERR688027', 'ERR688028', 'ERR688029',
                           'ERR688030', 'ERR688031', 'ERR688032', 'ERR688033', 'ERR688034', 'ERR688035', 'ERR688036',
                           'ERR688037', 'ERR688038', 'ERR688039', 'ERR688041', 'ERR688042', 'ERR688043', 'ERR688044',
                           'ERR688045', 'ERR688046', 'ERR688047', 'ERR688048', 'ERR688049', 'ERR688026', 'ERR688040']},
    #
    {
        'Source': "Genotyping of Mycobacterium tuberculosis spreading in Hanoi, Vietnam using conventional and whole genome sequencing methods",
        'Author': "S Maeda, M Hijikata, NT Le Hang, PH Thuong...(2020)",
        'study accession number': 'DRP005596',
        'location': 'Vietnam',
        'run accessions': ['DRR184599'] + ['DRR1846' + str(k).zfill(2) for k in range(83)] + ['DRR184' + str(k) for k in
                                                                                              range(877, 1000)] + [
                              'DRR185' + str(k).zfill(3) for k in range(125)]},
    # DRR184599–682 and DRR184877–5124
]

from xlrd import open_workbook

wwb = open_workbook('data/Brynildsrud_Dataset_S1.xls')
wws = wwb.sheet_by_index(0)

Brynildsrud = {}
for row in range(1, wws.nrows):
    srr = wws.cell_value(row, 4)
    if len(srr) > 1:
        Brynildsrud[srr] = {}
        Brynildsrud[srr]['Source'] = wws.cell_value(row, 5).replace(',', '')
        if Brynildsrud[srr]['Source'] == 'This study':
            Brynildsrud[srr]['Author'] = 'Brynildsrud et al.'
        else:
            Brynildsrud[srr]['Author'] = ''
        Brynildsrud[srr]['Source'] = Brynildsrud[srr]['Source'].replace('This study',
                                                                        'Global expansion of Mycobacterium tuberculosis lineage 4 shaped by colonial migration and local adaptation')
        Brynildsrud[srr]['study accession number'] = ''
        if len(wws.cell_value(row, 3)) > 1:
            Brynildsrud[srr]['study accession number'] = wws.cell_value(row, 3)
        Brynildsrud[srr]['location'] = ''
        if len(wws.cell_value(row, 6)) > 1:
            Brynildsrud[srr]['location'] = wws.cell_value(row, 6)
        Brynildsrud[srr]['date'] = ''
        if wws.cell_value(row, 0).count('_') == 2:
            dat = wws.cell_value(row, 0).split('_')[-1]
            unedate = True
            for w in dat:
                if w not in '0123456789':
                    unedate = False
            if unedate:
                Brynildsrud[srr]['date'] = dat


def get_info(srr):
    ret = Entrez.efetch(db="sra", id=srr, retmode="xml")
    dico = xmltodict.parse(ret.read())
    # On ne gère (pour l'instant ?) que les ILLUMINA paired end
    if 'ILLUMINA' in dico['EXPERIMENT_PACKAGE_SET']['EXPERIMENT_PACKAGE']['EXPERIMENT']['PLATFORM'] \
            and 'PAIRED' in \
            dico['EXPERIMENT_PACKAGE_SET']['EXPERIMENT_PACKAGE']['EXPERIMENT']['DESIGN']['LIBRARY_DESCRIPTOR'][
                'LIBRARY_LAYOUT']:
        # On récupère diverses informations
        try:
            attributes = dico['EXPERIMENT_PACKAGE_SET']['EXPERIMENT_PACKAGE']['SAMPLE']['SAMPLE_ATTRIBUTES'][
                'SAMPLE_ATTRIBUTE']
        except:
            return {}
        location, date, sra, center, strain = '', '', '', '', ''
        for k in attributes:
            if k['TAG'] == 'geographic location (country and/or sea)':
                location = k['VALUE']
            elif k['TAG'] == 'collection date':
                date = k['VALUE']
            elif k['TAG'] == 'SRA accession':
                sra = k['VALUE']
            elif k['TAG'] == 'INSDC center name':
                center = k['VALUE']
            elif k['TAG'] == 'Strain':
                strain = k['VALUE']
        print(dico['EXPERIMENT_PACKAGE_SET']['EXPERIMENT_PACKAGE']['STUDY']['IDENTIFIERS']['EXTERNAL_ID'])
        dico0 = {'location': location,
                 'date': date,
                 'SRA': sra,
                 'center': center,
                 'strain': strain,
                 'taxid': dico['EXPERIMENT_PACKAGE_SET']['EXPERIMENT_PACKAGE']['SAMPLE']['SAMPLE_NAME']['TAXON_ID'],
                 'name': dico['EXPERIMENT_PACKAGE_SET']['EXPERIMENT_PACKAGE']['SAMPLE']['SAMPLE_NAME'][
                     'SCIENTIFIC_NAME'],
                 'study': dico['EXPERIMENT_PACKAGE_SET']['EXPERIMENT_PACKAGE']['STUDY']['@alias']
                 }
        if isinstance(dico['EXPERIMENT_PACKAGE_SET']['EXPERIMENT_PACKAGE']['STUDY']['IDENTIFIERS']['EXTERNAL_ID'],
                      list):
            dico0['bioproject'] = \
                dico['EXPERIMENT_PACKAGE_SET']['EXPERIMENT_PACKAGE']['STUDY']['IDENTIFIERS']['EXTERNAL_ID'][0]['#text']
        else:
            dico0['bioproject'] = \
                dico['EXPERIMENT_PACKAGE_SET']['EXPERIMENT_PACKAGE']['STUDY']['IDENTIFIERS']['EXTERNAL_ID']['#text']

        dico0['center'] = ''
        try:
            dico0['center'] = dico['EXPERIMENT_PACKAGE_SET']['EXPERIMENT_PACKAGE']['EXPERIMENT']['@center_name']
        except:
            pass
    return dico0


# ok = []
# for k in dico_afr:
#     if 'strain' in dico_afr[k] and dico_afr[k]['strain'] != '':
#         if dico_afr[k]['strain'] == 'Thailand':
#             dico_afr[k]['strain'] = ''
#         else:
#             ok.append((dico_afr[k]['strain'], k))
# for o in sorted(ok):
#     print(' : '.join(o))

# dico_afr['SRR3647357']['strain']# = 'Mycobacterium tuberculosis variant caprae'
# save_dico()

import subprocess
from os import system, stat
from shutil import copyfile

avant_IS = []
taille_gen = len(h37Rv)

# Mycobacterium tuberculosis variant caprae ERR1462634
# Mycobacterium tuberculosis variant microti SRR3647357
# H37Rv ERR305600
Nouvel_article = ['ERR3335723', 'ERR3335724', 'ERR3335725', 'ERR3335726', 'ERR3335727', 'ERR3335728', 'ERR3335729',
                  'ERR3335730', 'ERR3335731', 'ERR3335732', 'ERR3335733', 'ERR3335734', 'ERR3335735', 'ERR3335736',
                  'ERR3335737', 'ERR3335738', 'ERR3335739', 'ERR3335740', 'ERR3335741', 'ERR3335742', 'ERR3335743',
                  'ERR3335744', 'ERR3335745', 'ERR3335746', 'ERR3335747', 'ERR3335748', 'ERR3335749', 'ERR3335750',
                  'ERR3335751', 'ERR3335752', 'ERR3335753', 'ERR3335754', 'ERR3335755', 'ERR3335756', 'ERR3335757',
                  'ERR3335758', 'ERR3335759', 'ERR3335760', 'ERR3335761', 'ERR3335762', 'ERR3335763', 'ERR3335764',
                  'ERR3335765', 'ERR3335766', 'ERR3335767', 'ERR3335768', 'ERR3335769', 'ERR3335770']
marinum = ['SRR8368696', 'SRR8368678', 'SRR8368689']
Momies = ['ERR650569', 'ERR650970', 'ERR650971', 'ERR650973', 'ERR650974', 'ERR650975', 'ERR650976', 'ERR650977',
          'ERR650978', 'ERR650979', 'ERR650980', 'ERR650981', 'ERR650982', 'ERR650983', 'ERR650984', 'ERR650985',
          'ERR650986', 'ERR650987', 'ERR650988', 'ERR650990', 'ERR650991', 'ERR650992', 'ERR650993', 'ERR650994',
          'ERR650995', 'ERR650996', 'ERR650997', 'ERR650998', 'ERR650999', 'ERR651000', 'ERR651001', 'ERR651002',
          'ERR651003', 'ERR651005', 'ERR651006', 'ERR651007', 'ERR651008', 'ERR651009', 'ERR651010', 'ERR650989',
          'ERR651004', 'ERR650972']
Manilla = ['DRR099684', 'DRR099686', 'DRR099689', 'DRR099692', 'DRR099683', 'DRR099685', 'DRR099687', 'DRR099688',
           'DRR099690', 'DRR099691', 'DRR099693', 'DRR099694']
for item in listdir(REP + 'sequences/'):
    if '.fasta' not in item:
        # if item not in dico_afr: # and item[0] == 'S': #E pour ERR (pour Christophe), à remplacer par S (SRR, pour Guislaine)
        print('\n\n' + item + ' ' + str(len(dico_afr) + 1) + "/" + str(len(listdir(REP + 'sequences/'))))
        # system('cp data/dico_africanum.pkl data/dico_africanum_old.pkl')
        if item not in dico_afr:
            dico_afr[item] = {}
        rep = REP + 'sequences/' + item + '/'
        if item not in listdir(REP + 'sequences/'):
            mkdir(rep)
        if 'low_cover.txt' in listdir(rep):
            print("       => trop faible couverture!")
            continue
        if 'dico.txt' in listdir(rep):
            dico_afr[item] = eval(open(rep + 'dico.txt').read())
            for cle in dico_afr[item]:
                if 'spoligo' not in cle:
                    print('   -', cle, ':', dico_afr[item][cle])
            for cle in dico_afr[item]:
                if 'spoligo' in cle and 'new' not in cle:
                    print('   -', cle, ':\n    ', dico_afr[item][cle])
            for cle in dico_afr[item]:
                if 'spoligo' in cle and 'new' in cle:
                    print('   -', cle, ':\n    ', dico_afr[item][cle])
            dico_afr[item]['SRA'] = item
            to_save = False
        if len([u for u in listdir(rep) if 'fasta' in u]) == 0:
            print("   - Téléchargement des fasta")
            completed = subprocess.run(
                ['parallel-fastq-dump', '-t', '8', '--split-files', '--fasta', '-O', REP, '-s', item])
            if completed.returncode == 0:
                for k in listdir(REP):
                    if k.endswith('.fasta'):
                        shutil.move(REP + k, rep + k)
            else:
                print("Erreur de téléchargement")
                del dico_afr[item]
                save_dico()
                continue
        if (item + '_1.fasta' not in listdir(rep) or item + '_2.fasta' not in listdir(
                rep)):  # or (item+'_1.fasta' not in listdir(rep) and item+'_3.fasta' not in listdir(rep)):
            print("   - Les fichiers fasta n'ont pas la bonne forme")
            del dico_afr[item]
            shutil.rmtree(rep)
            save_dico()
            continue
        if item + '_shuffled.fasta' not in listdir(rep):
            print(
                "   - On mélange les deux fichiers fasta ainsi téléchargés, correspondant aux deux extrémités des fragments")
            for fic in ['_1', '_2']:
                system("sed -i 's/" + item + './' + item + fic + "./g' " + rep + item + fic + '.fasta')
            system("cat " + rep + item + '_1.fasta ' + rep + item + '_2.fasta > ' + rep + item + '_shuffled.fasta')
            '''
            for fic in ['_1', '_2']:
                with open(rep+item+fic+'.fasta', 'r') as f:
                    txt = f.read()
                txt = txt.replace(item+'.', item+fic+'.')
                with open(rep+item+fic+'.fasta', 'w') as f:
                    f.write(txt)
            with open(rep+item+'_shuffled.fasta', 'w') as f:
                f.write(open(rep+item+'_1.fasta', 'r').read())
                f.write(open(rep+item+'_2.fasta', 'r').read())
            '''
        if 'nb_reads' not in dico_afr[item] or dico_afr[item]['nb_reads'] == '':
            system('cat ' + rep + item + "_shuffled.fasta | grep '>' | wc -l > /tmp/nb.txt")
            nb = eval(open('/tmp/nb.txt').read().split('\n')[0])
            # nb = open(rep+item+'_shuffled.fasta').read().count('>')
            print("   - Nombre de reads :", nb)
            dico_afr[item]['nb_reads'] = nb
        if 'len_reads' not in dico_afr[item]:
            nb = len(''.join(open(rep + item + '_shuffled.fasta').read(10000).split('>')[1].split('\n')[1:]))
            print("   - Longueur des reads :", nb)
            dico_afr[item]['len_reads'] = nb
        if 'couverture' not in dico_afr[item] or dico_afr[item]['couverture'] == '':
            dico_afr[item]['couverture'] = round(dico_afr[item]['nb_reads'] * dico_afr[item]['len_reads'] / taille_gen,
                                                 2)
            print("   - Couverture :", dico_afr[item]['couverture'])
        if dico_afr[item]['couverture'] < 50:
            print("       => trop faible couverture!")
            system('touch ' + rep + 'low_cover.txt')
            del dico_afr[item]
            save_dico()
        else:
            if item + '.nal' not in listdir(rep) and item + '.nin' not in listdir(rep):
                print("   - On fait une BDD pour blast")
                completed = subprocess.run(
                    ['makeblastdb', '-in', rep + item + '_shuffled.fasta', '-dbtype', 'nucl', '-title', item, '-out',
                     rep + item])
                assert completed.returncode == 0
            if True:  # 'Source' not in dico_afr[item]:
                for u in Origines:
                    if item in u['run accessions']:
                        for uu in ['Source', 'Author', 'study accession number', 'location']:
                            try:
                                dico_afr[item][uu] = u[uu]
                                print("   - " + uu + ' : ' + dico_afr[item][uu])
                            except:
                                pass
            if 'taxid' not in dico_afr[item]:
                dicobis = get_info(item)
                for uu in dicobis:
                    dico_afr[item][uu] = dicobis[uu]
                    print("   - " + uu + ' : ' + dico_afr[item][uu])
            if item in Brynildsrud:
                for uu in Brynildsrud[item]:
                    dico_afr[item][uu] = Brynildsrud[item][uu]
                    print("   - " + uu + ' : ' + dico_afr[item][uu])
            if 'spoligo' not in dico_afr[item] or dico_afr[item]['spoligo'] == '':
                print("   - On blaste les spoligos")
                dico_afr[item]['spoligo'] = ''
                dico_afr[item]['spoligo_new'] = ''
                completed = subprocess.run(
                    "blastn -num_threads 12 -query data/spoligo_old.fasta -evalue 1e-6 -task blastn -db " + rep + item + " -outfmt '10 qseqid sseqid sstart send qlen length score evalue' -out /tmp/" + item + "_old.blast",
                    shell=True)
                assert completed.returncode == 0
                completed = subprocess.run(
                    "blastn -num_threads 12 -query data/spoligo_new.fasta -evalue 1e-6 -task blastn -db " + rep + item + " -outfmt '10 qseqid sseqid sstart send qlen length score evalue' -out /tmp/" + item + "_new.blast",
                    shell=True)
                assert completed.returncode == 0
                print("   - On écrit les spoligos obtenus dans le fichier csv")
                for pos, spol in enumerate(['old', 'new']):
                    with open('/tmp/' + item + '_' + spol + '.blast') as f:
                        matches = f.read()
                        nb = open('data/spoligo_' + spol + '.fasta').read().count('>')
                        for k in range(1, nb + 1):
                            # if 'espaceur'+spol.capitalize()+str(k) in matches:
                            # if matches.count('espaceur'+spol.capitalize()+str(k)+',')/dico_afr[item]['couverture']>0.05:
                            if matches.count('espaceur' + spol.capitalize() + str(k) + ',') >= 5:
                                dico_afr[item]['spoligo' + ['', '_new'][pos]] += '\u25A0'
                            else:
                                dico_afr[item]['spoligo' + ['', '_new'][pos]] += '\u25A1'
                    dico_afr[item]['spoligo' + ['', '_new'][pos] + '_nb'] = [
                        matches.count('espaceur' + spol.capitalize() + str(k) + ',') for k in range(1, nb + 1)]
                    system('mv /tmp/' + item + '_' + spol + '.blast ' + rep)
                # print("     Lignée : "+", ".join(dico_afr[item]['lineage_Coll']))
                save_dico()
                print("     " + dico_afr[item]['spoligo'])
                print("     " + str(dico_afr[item]['spoligo_nb']))
                print("     " + dico_afr[item]['spoligo_new'])
                print("     " + str(dico_afr[item]['spoligo_new_nb']))
            if 'spoligo_vitro' not in dico_afr[item]:
                print("   - On blaste les spoligos in vitro")
                evalue = '1e-6'
                dico_afr[item]['spoligo_vitro'] = ''
                dico_afr[item]['spoligo_vitro_new'] = ''
                completed = subprocess.run(
                    "blastn -num_threads 8 -query data/spoligo_vitro.fasta -evalue " + evalue + " -task blastn -db " + rep + item + " -outfmt '10 qseqid sseqid sstart send qlen length score evalue' -out /tmp/" + item + "_vitro.blast",
                    shell=True)
                assert completed.returncode == 0
                completed = subprocess.run(
                    "blastn -num_threads 8 -query data/spoligo_vitro_new.fasta -evalue " + evalue + " -task blastn -db " + rep + item + " -outfmt '10 qseqid sseqid sstart send qlen length score evalue' -out /tmp/" + item + "_vitro_new.blast",
                    shell=True)
                assert completed.returncode == 0
                # print("   - On écrit les spoligos vitro obtenus dans le fichier csv")
                with open('/tmp/' + item + '_vitro.blast') as f:
                    matches = f.read()
                    nb = int(open('data/spoligo_vitro.fasta').read().count('>') / 2)
                    for k in range(1, nb + 1):
                        # if 'espaceur'+spol.capitalize()+str(k) in matches:
                        if min([matches.count('espaceur_vitroOld' + str(k) + ','),
                                matches.count('espaceur_vitroBOld' + str(k) + ',')]) / dico_afr[item][
                            'couverture'] > 0.05:
                            dico_afr[item]['spoligo_vitro'] += '\u25A0'
                        else:
                            dico_afr[item]['spoligo_vitro'] += '\u25A1'
                dico_afr[item]['spoligo_vitro_nb'] = [(matches.count('espaceur_vitroOld' + str(k) + ','),
                                                       matches.count('espaceur_vitroBOld' + str(k) + ',')) for k in
                                                      range(1, nb + 1)]
                with open('/tmp/' + item + '_vitro_new.blast') as f:
                    matches = f.read()
                    nb = int(open('data/spoligo_vitro_new.fasta').read().count('>') / 2)
                    for k in range(1, nb + 1):
                        if min([matches.count('espaceur_vitro_new' + str(k) + ','),
                                matches.count('espaceur_vitro_newB' + str(k) + ',')]) / dico_afr[item][
                            'couverture'] > 0.05:
                            dico_afr[item]['spoligo_vitro_new'] += '\u25A0'
                        else:
                            dico_afr[item]['spoligo_vitro_new'] += '\u25A1'
                print("     " + dico_afr[item]['spoligo_vitro'])
                print("     " + dico_afr[item]['spoligo_vitro_new'])
                dico_afr[item]['spoligo_vitro_new_nb'] = [(matches.count('espaceur_vitro_new' + str(k) + ','),
                                                           matches.count('espaceur_vitro_newB' + str(k) + ',')) for k in
                                                          range(1, nb + 1)]
                system('mv /tmp/' + item + '_*.blast ' + rep)
                print("     " + str(dico_afr[item]['spoligo_vitro_nb']))
                print("     " + str(dico_afr[item]['spoligo_vitro_new_nb']))
                save_dico()
            if 'SIT' not in dico_afr[item] or dico_afr[item]['SIT'] == '':
                spol = dico_afr[item]['spoligo']
                # spol = ''.join(list(map(change, spol)))
                if spol in spol_sit:
                    dico_afr[item]['SIT'] = spol_sit[spol]
                else:
                    dico_afr[item]['SIT'] = 'X'
                print("   - On ajoute le SIT :", dico_afr[item]['SIT'])
                save_dico()
            if 'SIT_silico' not in dico_afr[item]:
                spol = dico_afr[item]['spoligo_vitro']
                # spol = ''.join(list(map(change, spol)))
                if spol in spol_sit:
                    dico_afr[item]['SIT_silico'] = spol_sit[spol]
                else:
                    dico_afr[item]['SIT_silico'] = 'X'
                print("   - On ajoute le SIT silico :", dico_afr[item]['SIT_silico'])
                save_dico()
            if 'lineage_Coll' not in dico_afr[item] or dico_afr[item]['lineage_Coll'] == '':
                print("   - On ajoute la lignée selon les SNPs Coll")
                lignee = []
                for row in ws.iter_rows(min_row=2):
                    if row[1].value != None:
                        pos = row[1].value - 1
                        assert h37Rv[pos] == row[3].value.split('/')[0]
                        # print(row[0].value, row[1].value-1, row[3].value.split('/'))
                        seq1 = h37Rv[pos - longueur:pos + longueur + 1]
                        if '*' not in row[0].value:
                            seq2 = seq1[:20] + row[3].value.split('/')[1] + seq1[21:]
                        else:
                            seq1 = seq1[:20] + row[3].value.split('/')[1] + seq1[21:]
                            seq2 = seq1[:20] + row[3].value.split('/')[0] + seq1[21:]
                        with open('/tmp/snp.fasta', 'w') as f:
                            f.write('>\n' + seq2)
                        result = subprocess.run(
                            ["blastn", "-num_threads", "8", "-query", "/tmp/snp.fasta", "-evalue", "1e-5", "-task",
                             "blastn", "-db", rep + item, "-outfmt", "10 sseq"], stdout=subprocess.PIPE)
                        formated_results = result.stdout.decode('utf8').splitlines()
                        # nb_seq1 = formated_results.count(seq1)
                        # nb_seq2 = formated_results.count(seq2)
                        nb_seq1 = len([u for u in formated_results if seq1[20:25] in u]) + len(
                            [u for u in formated_results if seq1[16:21] in u])
                        # nb_seq2 = formated_results.count(seq2)
                        nb_seq2 = len([u for u in formated_results if seq2[20:25] in u]) + len(
                            [u for u in formated_results if seq2[16:21] in u])
                        # print(row[0].value.replace('lineage', '').replace('*',''),nb_seq1,nb_seq2)
                        if nb_seq2 > nb_seq1:
                            lignee.append(row[0].value.replace('lineage', '').replace('*', ''))
                        '''
                        a=formated_results.count(seq1[:20]+row[3].value.split('/')[0]+seq1[21:])
                        b=formated_results.count(seq1[:20]+row[3].value.split('/')[1]+seq1[21:])
                        c=formated_results.count(seq1)
                        d=formated_results.count(h37Rv[pos-longueur:pos+longueur+1])
                        if row[0].value.replace('lineage', '').replace('*','') not in ['4', '4.9']:
                            if nb_seq2*0.1>nb_seq1 or (nb_seq2>0 and nb_seq1==0):
                                lignee.append(row[0].value.replace('lineage', '').replace('*',''))
                            if d==0 and b > 0: 
                                lignee.append(row[0].value.replace('lineage', '').replace('*',''))
                        else:
                            if b==0 and d > 0:
                                lignee.append(row[0].value.replace('lineage', '').replace('*',''))  '''
                lignee = sorted(set(lignee))
                dico_afr[item]['lineage_Coll'] = lignee
                print("     Lignée (Coll) : " + ", ".join(dico_afr[item]['lineage_Coll']))
                save_dico()
            if 'lineage_L6+animal' not in dico_afr[item]:
                print("   - On ajoute la lignée selon les SNPs L6+animal")
                seq1 = 'ACGTCGATGGTCGCGACCTCCGCGGCATAGTCGAA'
                seq2 = "ACGTCGATGGTCGCGACTTCCGCGGCATAGTCGAA"
                with open('/tmp/snp.fasta', 'w') as f:
                    f.write('>\n' + seq2)
                result = subprocess.run(
                    ["blastn", "-num_threads", "12", "-query", "/tmp/snp.fasta", "-evalue", "1e-5", "-task", "blastn",
                     "-db", rep + item, "-outfmt", "10 sseq"], stdout=subprocess.PIPE)
                formated_results = result.stdout.decode('utf8').splitlines()
                nb_seq1 = len([u for u in formated_results if seq1[13:18] in u]) + len(
                    [u for u in formated_results if seq1[17:22] in u])
                nb_seq2 = len([u for u in formated_results if seq2[13:18] in u]) + len(
                    [u for u in formated_results if seq2[17:22] in u])
                if nb_seq1 > nb_seq2:
                    dico_afr[item]['lineage_L6+animal'] = '1'
                elif nb_seq2 > nb_seq1:
                    dico_afr[item]['lineage_L6+animal'] = '2'
                else:
                    dico_afr[item]['lineage_L6+animal'] = 'X'
                print("     Lignée (L6+animal) : " + dico_afr[item]['lineage_L6+animal'])
                save_dico()
            if 'lineage_PGG' not in dico_afr[item]:
                print("   - On ajoute la lignée selon les SNPs PGG")
                lineage = []
                pos = 2154724
                seq1 = h37Rv[pos - longueur:pos + longueur + 1]
                seq2 = seq1[:19] + 'A' + seq1[20:]
                with open('/tmp/snp.fasta', 'w') as f:
                    f.write('>\n' + seq2)
                result = subprocess.run(
                    ["blastn", "-num_threads", "12", "-query", "/tmp/snp.fasta", "-evalue", "1e-5", "-task", "blastn",
                     "-db", rep + item, "-outfmt", "10 sseq"], stdout=subprocess.PIPE)
                formated_results = result.stdout.decode('utf8').splitlines()
                # print(formated_results)
                # nb_seq1 = formated_results.count(seq1)
                nb_seq1 = len([u for u in formated_results if seq1[19:24] in u]) + len(
                    [u for u in formated_results if seq1[15:20] in u])
                # nb_seq2 = formated_results.count(seq2)
                nb_seq2 = len([u for u in formated_results if seq2[19:24] in u]) + len(
                    [u for u in formated_results if seq2[15:20] in u])
                if nb_seq1 > nb_seq2:
                    lineage.append('2')
                elif nb_seq2 > nb_seq1:
                    lineage.append('1')
                else:
                    lineage.append('X')
                pos = 7585 - 1
                seq1 = h37Rv[pos - longueur:pos + longueur + 1]
                seq2 = seq1[:20] + 'C' + seq1[21:]
                with open('/tmp/snp.fasta', 'w') as f:
                    f.write('>\n' + seq2)
                result = subprocess.run(
                    ["blastn", "-num_threads", "12", "-query", "/tmp/snp.fasta", "-evalue", "1e-5", "-task", "blastn",
                     "-db", rep + item, "-outfmt", "10 sseq"], stdout=subprocess.PIPE)
                formated_results = result.stdout.decode('utf8').splitlines()
                # nb_seq1 = formated_results.count(seq1)
                nb_seq1 = len([u for u in formated_results if seq1[20:25] in u]) + len(
                    [u for u in formated_results if seq1[16:21] in u])
                # nb_seq2 = formated_results.count(seq2)
                nb_seq2 = len([u for u in formated_results if seq2[20:25] in u]) + len(
                    [u for u in formated_results if seq2[16:21] in u])
                if nb_seq1 > nb_seq2:
                    lineage.append('3')
                elif nb_seq2 > nb_seq1:
                    lineage.append('1')
                else:
                    lineage.append('X')
                dico_afr[item]['lineage_PGG_cp'] = lineage
                if lineage == ['1', '1']:
                    dico_afr[item]['lineage_PGG'] = '1'
                elif lineage in [['1', '2'], ['2', '1']]:
                    dico_afr[item]['lineage_PGG'] = '2'
                elif lineage in [['2', '3'], ['3', '2']]:
                    dico_afr[item]['lineage_PGG'] = '3'
                else:
                    dico_afr[item]['lineage_PGG'] = 'X'
                print("     Lignée (PGG) : " + dico_afr[item]['lineage_PGG'] + ' (' + ", ".join(
                    dico_afr[item]['lineage_PGG_cp']) + ')')
                save_dico()
            if 'lineage_Pali' not in dico_afr[item]:
                print("   - On ajoute la lignée selon les SNPs Palittapongarnpim")
                lignee = []
                cpt = 1
                for item2, pos0 in enumerate(Lignee_Pali):
                    seq1, seq2 = Lignee_Pali[pos0][:2]
                    with open('/tmp/snp.fasta', 'w') as f:
                        f.write('>\n' + seq2)
                    cmd = "blastn -query /tmp/snp.fasta -num_threads 12 -evalue 1e-5 -task blastn -db " + rep + item + " -outfmt '10 sseq' -out /tmp/snp_Pali.blast"
                    system(cmd)
                    with open("/tmp/snp_Pali.blast") as f:
                        formated_results = f.read().splitlines()
                    # nb_seq1 = formated_results.count(seq1)
                    # nb_seq2 = formated_results.count(seq2)
                    nb_seq1 = len([u for u in formated_results if seq1[20:25] in u]) + len(
                        [u for u in formated_results if seq1[16:21] in u])
                    # nb_seq2 = formated_results.count(seq2)
                    nb_seq2 = len([u for u in formated_results if seq2[20:25] in u]) + len(
                        [u for u in formated_results if seq2[16:21] in u])

                    if nb_seq2 > nb_seq1:  # or (nb_seq2>0 and nb_seq1==0):
                        cpt += 1
                        lignee.append(Lignee_Pali[pos0][2])
                lignee = [u for u in sorted(set(lignee))]
                dico_afr[item]['lineage_Pali'] = lignee
                print("     Lignée (Pali) : " + ", ".join(dico_afr[item]['lineage_Pali']))
                save_dico()
            if 'lineage_Shitikov' not in dico_afr[item]:
                print("   - On ajoute la lignée selon les SNPs Shitikov")
                lignee = []
                cpt = 1
                for item2, pos0 in enumerate(Lignee_Shitikov):
                    seq1, seq2 = Lignee_Shitikov[pos0][:2]
                    with open('/tmp/snp.fasta', 'w') as f:
                        f.write('>\n' + seq2)
                    cmd = "blastn -query /tmp/snp.fasta -num_threads 12 -evalue 1e-5 -task blastn -db " + rep + item + " -outfmt '10 sseq' -out /tmp/snp_Shitikov.blast"
                    system(cmd)
                    with open("/tmp/snp_Shitikov.blast") as f:
                        formated_results = f.read().splitlines()
                    # nb_seq1 = formated_results.count(seq1)
                    # nb_seq2 = formated_results.count(seq2)
                    nb_seq1 = len([u for u in formated_results if seq1[20:25] in u]) + len(
                        [u for u in formated_results if seq1[16:21] in u])
                    # nb_seq2 = formated_results.count(seq2)
                    nb_seq2 = len([u for u in formated_results if seq2[20:25] in u]) + len(
                        [u for u in formated_results if seq2[16:21] in u])

                    if nb_seq2 > nb_seq1:  # or (nb_seq2>0 and nb_seq1==0):
                        cpt += 1
                        lignee.append(Lignee_Shitikov[pos0][2])
                lignee = [u for u in sorted(set(lignee))]
                dico_afr[item]['lineage_Shitikov'] = lignee
                print("     Lignée (Shitikov) : " + ", ".join(dico_afr[item]['lineage_Shitikov']))
                save_dico()
            if 'Lignee_Stucki' not in dico_afr[item]:
                print("   - On ajoute la lignée selon les SNPs Stucki")
                lignee = []
                cpt = 1
                for item2, pos0 in enumerate(Lignee_Stucki):
                    seq1, seq2 = Lignee_Stucki[pos0][:2]
                    with open('/tmp/snp.fasta', 'w') as f:
                        f.write('>\n' + seq2)
                    cmd = "blastn -query /tmp/snp.fasta -num_threads 12 -evalue 1e-5 -task blastn -db " + rep + item + " -outfmt '10 sseq' -out /tmp/snp_Stucki.blast"
                    system(cmd)
                    with open("/tmp/snp_Stucki.blast") as f:
                        formated_results = f.read().splitlines()
                    # nb_seq1 = formated_results.count(seq1)
                    # nb_seq2 = formated_results.count(seq2)
                    nb_seq1 = len([u for u in formated_results if seq1[20:25] in u]) + len(
                        [u for u in formated_results if seq1[16:21] in u])
                    # nb_seq2 = formated_results.count(seq2)
                    nb_seq2 = len([u for u in formated_results if seq2[20:25] in u]) + len(
                        [u for u in formated_results if seq2[16:21] in u])
                    # print(Lignee_Stucki[pos0][2],nb_seq1,nb_seq2)
                    if nb_seq2 > nb_seq1:  # or (nb_seq2>0 and nb_seq1==0):
                        cpt += 1
                        lignee.append(Lignee_Stucki[pos0][2])
                lignee = [u for u in sorted(set(lignee))]
                if '4.10' in lignee:
                    lignee.remove('4.10')
                else:
                    lignee.append('4.10')
                dico_afr[item]['Lignee_Stucki'] = lignee
                print("     Lignée (Stucki) : " + ", ".join(dico_afr[item]['Lignee_Stucki']))
                save_dico()
                '''if dico_afr[item]['IS_mapper'] == '':
                if 'ISmapper' not in listdir(REP) or item not in listdir(REP+'ISmapper/'):
                    print("   - On exécute ISmapper")    
                    rename(rep+item+'_1.fasta', rep+item+'_1.fastq')
                    rename(rep+item+'_2.fasta', rep+item+'_2.fastq')
                    completed = subprocess.run('ismap --reads '+rep+'*fastq --queries '+REP+'data/IS6110.fasta --reference '+REP+'data/H37Rv.gb --output_dir '+REP+'ISmapper', shell = True)
                    assert completed.returncode == 0
                    rename(rep+item+'_1.fastq', rep+item+'_1.fasta')
                    rename(rep+item+'_2.fastq', rep+item+'_2.fasta')
                    for fic in listdir(REP):
                        if '.log' in fic:
                            rename(REP+fic, REP+'ISmapper/'+item+'/IS6110/'+fic)
                print("   - On collecte les résultats d'ISmapper")
                dico_afr[item]['IS_mapper'] = open(REP+'ISmapper/'+item+'/IS6110/'+item+'__NC_000962.3_table.txt').read().split('\n')[1]
                save_dico()'''
                '''
            if dico_afr[item]['MIRU02'] == '':
                print("   - On exécute MIRU-profiler")    
                system(REP+'./MIRU-profiler -i '+rep+item+'_shuffled.fasta')
                # On inscrit les MIRU dans le dico
                res=open(REP+'24MIRU_results').read().split('\n')
                if len(res)>1:
                    res = [u for u in res if len(u)>1]
                    for item2, k in enumerate(res[0].split('\t')[1:]):
                        dico_afr[item][k] = res[-1].split('\t')[item2+1]
                    save_dico()
                # On effectue du nettoyage
                rename(REP+'24MIRU_results', rep+'24MIRU_results')
                shutil.rmtree(REP+'OUTPUT')'''
                """
            if dico_afr[item]['MIRU02'] == '':
                if  'auto' not in ''.join(listdir(rep)):
                    print("   - On exécute velveth")
                    #system(REP+'./velveth '+rep+'auto 23,52,2 -fasta -shortPaired1 '+rep+item+'_shuffled.fasta')
                '''dicoN50 = {}
                for nom in listdir(rep):
                    if 'auto' in nom:
                        with open(rep+nom+'/Log') as f:
                            log = f.read()
                            if 'n50' in log:
                                dicoN50[nom] = eval(log.split('n50 of ')[1].split(',')[0])
                            else:
                                print("   - On exécute velvetg")    
                                system(REP+'./velvetg '+rep+nom+' -exp_cov auto')                            
                                dicoN50[nom] = eval(log.split('n50 of ')[1].split(',')[0])
                print("   - On ne conserve que le meilleur contig")    
                best_contig = max([(dicoN50[cle], cle) for cle in dicoN50])[1]
                for nom in listdir(rep):
                    if 'auto' in nom and nom != best_contig:
                        shutil.rmtree(rep+nom)
                print("   - On exécute MIRU-profiler")    
                system(REP+'./MIRU-profiler -i '+rep+best_contig+'/contigs.fa')
                # On inscrit les MIRU dans le dico
                res=open(REP+'24MIRU_results').read().split('\n')
                if len(res)>1:
                    res = [u for u in res if len(u)>1]
                    for item2, k in enumerate(res[0].split('\t')[1:]):
                        dico_afr[item][k] = res[-1].split('\t')[item2+1]
                    save_dico()
                # On effectue du nettoyage
                rename(REP+'24MIRU_results', rep+'24MIRU_results')
                shutil.rmtree(REP+'OUTPUT')'''
                """
            """if 'IS6110.log' not in listdir(rep) and item+"_IS.blast" not in listdir(rep): # or stat(rep+'IS6110.log').st_size == 0:
                print("   - On compte le nombre d'IS6110")
                cmd = "blastn -query data/IS6110.fasta -evalue 1e-5 -task blastn -db "+rep+item+" -max_target_seqs 2000000 -outfmt '10 qstart sstart send sseqid' -out /tmp/"+item+"_IS.blast"
                system(cmd)
                with open("/tmp/"+item+"_IS.blast") as f:
                    dd = f.read().splitlines()
                copyfile("/tmp/"+item+"_IS.blast", rep+item+"_IS.blast")
                ee=[]
                for k in range(41,70):
                    ee.extend([u for u in dd if u.startswith('1,'+str(k)+',')])
                ee = [(u.split(',')[1], u.split(',')[3]) for u in ee if eval(u.split(',')[1])<eval(u.split(',')[2])]
                if ee != []:
                    fasta_sequences = {}
                    for u in ['_1', '_2']:
                        fasta_sequences[item+u] = SeqIO.parse(open(rep+item+u+'.fasta'),'fasta')
                    avant = []
                    avant2 = []
                    for sens in fasta_sequences:
                        for fasta in fasta_sequences[sens]:
                            name, sequence = fasta.id, str(fasta.seq)
                            for u in ee:
                                if name == u[1]:
                                    seq = sequence[eval(u[0])-41:eval(u[0])-1]
                                    avant.append(seq)
                                    avant2.append(sequence[eval(u[0])-41:eval(u[0])+41])
                                    '''
                                    if seq not in avant_sequences:
                                        avant_sequences.append(seq)
                                    if avant_IS == [] or  max([similaire(seq, v) for v in avant_IS]) < 0.7:
                                        avant_IS.append(seq)
                                        quel_IS = seq
                                    else:
                                        _, quel_IS = max([(similaire(seq, v), v) for v in avant_IS], key=lambda x:x[0])
                                    if quel_IS+'_IS' not in dico_afr[item]:
                                        for item2 in list(dico_afr):
                                            dico_afr[item2][quel_IS+'_IS'] = 0
                                    dico_afr[item][quel_IS+'_IS'] += 1
                                    if avant == []:
                                        avant = [seq]
                                    elif seq not in avant and max([similaire(seq, v) for v in avant]) < 0.7:
                                        avant.append(seq)'''
                    #print(avant)
                    #dico_afr[item]['IS6110'] = str(len(avant))
                    with open(rep+'IS6110.log','w') as f:
                        f.write('\n'.join(avant))
                    with open(rep+'IS6110_2.log','w') as f:
                        f.write('\n'.join(avant2))
                    save_dico()
            """
        for k in Origines:
            if item in k['run accessions']:
                if 'location' in k:
                    dico_afr[item]['location'] = k['location']
        if item in dico_afr:
            if dico_afr[item].get('REP', '') == '':
                dico_afr[item]['REP'] = REP
            # On rajoute d'éventuelles clé manquantes:
            if 'name' not in dico_afr[item]:
                dico_afr[item]['name'] = ''
            # On supprime d'éventuels métagénomes, etc.
            if 'low_cover.txt' not in listdir(rep):
                with open(rep + 'dico.txt', 'w') as f:
                    f.write(str(dico_afr[item]))
                '''with open('data/spoligos_me.csv', mode='w') as csv_file:
                    fieldnames = list(dico_afr[next(iter(dico_afr))].keys())
                    writer = csv.DictWriter(csv_file, fieldnames=fieldnames)#, 'name', 'accession', 'date', 'taxid', 'IS6110', 'location', 'center', 'strain'])
                    writer.writeheader()
                    for k in dict(dico_afr):
                        if dico_afr[k] != {}:
                            writer.writerow(dico_afr[k])
                with open('data/spoligos.csv', 'w') as f1:
                    with open('data/spoligos_me.csv', 'r') as f2:
                        f1.write(f2.read().replace('\u25A0', 'n').replace('\u25A1', 'o'))'''
            if 'metagenome' in dico_afr[item]['name'] and item in listdir(REP + 'sequences'):
                print("    => Il s'agit d'un métagénome, on le supprime")
                del dico_afr[item]
                try:
                    system('rm -fr ' + rep)
                except:
                    print("     -> Pas trouvé")
        save_dico()

for k in dico_afr:
    if 'Mycobacterium' not in dico_afr[k]['name'] and len(dico_afr[k]['name']) > 0:
        print(dico_afr[k]['name'])
